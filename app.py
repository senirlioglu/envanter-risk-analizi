import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import zipfile
from supabase import create_client, Client

# Mobil uyumlu sayfa ayarı
st.set_page_config(page_title="Envanter Risk Analizi", layout="wide", page_icon="📊")

# ==================== SUPABASE BAĞLANTISI ====================
SUPABASE_URL = "https://tlcgcdiycgfxpxwzkwuf.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRsY2djZGl5Y2dmeHB4d3prd3VmIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjU2NDgwMjksImV4cCI6MjA4MTIyNDAyOX0.4GnWTvUmdLzqcP0v8MAqaNUQkYgk0S8qrw6nSPsz-t4"

@st.cache_resource
def get_supabase_client():
    return create_client(SUPABASE_URL, SUPABASE_KEY)

supabase: Client = get_supabase_client()

# ==================== GİRİŞ SİSTEMİ ====================
USERS = {
    "ziya": "Gm2025!",
    "sm1": "Sm12025!",
    "sm2": "Sm22025!",
    "sm3": "Sm32025!",
    "sm4": "Sm42025!",
    "sma": "Sma2025!",
}

def login():
    if "user" not in st.session_state:
        st.session_state.user = None
    
    if st.session_state.user is None:
        st.markdown("""
        <div style="max-width: 400px; margin: 100px auto; padding: 40px; 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    border-radius: 15px; text-align: center;">
            <h1 style="color: white;">📊 Envanter Risk Analizi</h1>
            <p style="color: #eee;">Mağaza Detay Analizi</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            st.markdown("### 🔐 Giriş Yap")
            username = st.text_input("Kullanıcı Adı", key="login_user")
            password = st.text_input("Şifre", type="password", key="login_pass")
            
            if st.button("Giriş", use_container_width=True):
                if username.lower() in USERS and USERS[username.lower()] == password:
                    st.session_state.user = username.lower()
                    st.rerun()
                else:
                    st.error("❌ Hatalı kullanıcı adı veya şifre")
        st.stop()

login()

# ==================== SUPABASE FONKSİYONLARI ====================

def save_to_supabase(df_original):
    """Excel verisini Supabase'e kaydet (duplicate kontrolü ile)"""
    try:
        df = df_original.copy()
        
        # Gerekli sütunlar var mı kontrol
        required_cols = ['Mağaza Kodu', 'Depolama Koşulu Grubu', 'Envanter Dönemi', 'Malzeme Kodu']
        for col in required_cols:
            if col not in df.columns:
                return 0, 0, f"'{col}' sütunu bulunamadı"
        
        # Unique key oluştur - Bu yükleme için hangi envanterler var
        df['_inv_key'] = df['Mağaza Kodu'].astype(str) + '_' + \
                         df['Depolama Koşulu Grubu'].astype(str) + '_' + \
                         df['Envanter Dönemi'].astype(str)
        
        unique_inventories = df['_inv_key'].unique().tolist()
        
        # Supabase'de bu envanterler var mı kontrol et
        existing_inventories = set()
        for inv_key in unique_inventories:
            parts = inv_key.split('_')
            if len(parts) >= 3:
                mag_kodu = parts[0]
                dep_grubu = '_'.join(parts[1:-1])  # Orta kısım (boşluk içerebilir)
                donem = parts[-1]
                
                try:
                    result = supabase.table('envanter_veri').select('id').eq(
                        'magaza_kodu', mag_kodu
                    ).eq(
                        'depolama_kosulu_grubu', dep_grubu
                    ).eq(
                        'envanter_donemi', donem
                    ).limit(1).execute()
                    
                    if result.data and len(result.data) > 0:
                        existing_inventories.add(inv_key)
                except:
                    pass
        
        # Sadece yeni envanterler
        new_inventories = [inv for inv in unique_inventories if inv not in existing_inventories]
        skipped_inventories = [inv for inv in unique_inventories if inv in existing_inventories]
        
        if not new_inventories:
            return 0, len(skipped_inventories), "Tüm envanterler zaten mevcut"
        
        # Sadece yeni envanterlerin verilerini filtrele
        df_new = df[df['_inv_key'].isin(new_inventories)].copy()
        
        # Sütun mapping
        col_mapping = {
            'Mağaza Kodu': 'magaza_kodu',
            'Mağaza Tanım': 'magaza_tanim',
            'Satış Müdürü': 'satis_muduru',
            'Bölge Sorumlusu': 'bolge_sorumlusu',
            'Depolama Koşulu Grubu': 'depolama_kosulu_grubu',
            'Depolama Koşulu': 'depolama_kosulu',
            'Envanter Dönemi': 'envanter_donemi',
            'Envanter Tarihi': 'envanter_tarihi',
            'Envanter Başlangıç Tarihi': 'envanter_baslangic_tarihi',
            'Ürün Grubu Kodu': 'urun_grubu_kodu',
            'Ürün Grubu Tanımı': 'urun_grubu_tanimi',
            'Mal Grubu Kodu': 'mal_grubu_kodu',
            'Mal Grubu Tanımı': 'mal_grubu_tanimi',
            'Malzeme Kodu': 'malzeme_kodu',
            'Malzeme Tanımı': 'malzeme_tanimi',
            'Satış Fiyatı': 'satis_fiyati',
            'Sayım Miktarı': 'sayim_miktari',
            'Sayım Tutarı': 'sayim_tutari',
            'Kaydi Miktar': 'kaydi_miktar',
            'Kaydi Tutar': 'kaydi_tutar',
            'Fark Miktarı': 'fark_miktari',
            'Fark Tutarı': 'fark_tutari',
            'Kısmi Envanter Miktarı': 'kismi_envanter_miktari',
            'Kısmi Envanter Tutarı': 'kismi_envanter_tutari',
            'Fire Miktarı': 'fire_miktari',
            'Fire Tutarı': 'fire_tutari',
            'Önceki Fark Miktarı': 'onceki_fark_miktari',
            'Önceki Fark Tutarı': 'onceki_fark_tutari',
            'Önceki Fire Miktarı': 'onceki_fire_miktari',
            'Önceki Fire Tutarı': 'onceki_fire_tutari',
            'Satış Miktarı': 'satis_miktari',
            'Satış Hasılatı': 'satis_hasilati',
            'İade Miktarı': 'iade_miktari',
            'İade Tutarı': 'iade_tutari',
            'İptal Fişteki Miktar': 'iptal_fisteki_miktar',
            'İptal Fiş Tutarı': 'iptal_fis_tutari',
            'İptal GP Miktarı': 'iptal_gp_miktari',
            'İptal GP Tutarı': 'iptal_gp_tutari',
            'İptal Satır Miktarı': 'iptal_satir_miktari',
            'İptal Satır Tutarı': 'iptal_satir_tutari',
        }
        
        # Veriyi hazırla
        records = []
        for _, row in df_new.iterrows():
            record = {}
            for excel_col, db_col in col_mapping.items():
                if excel_col in row.index:
                    val = row[excel_col]
                    if pd.isna(val):
                        val = None
                    elif isinstance(val, pd.Timestamp):
                        val = val.strftime('%Y-%m-%d')
                    elif isinstance(val, (np.integer, np.int64)):
                        val = int(val)
                    elif isinstance(val, (np.floating, np.float64)):
                        val = float(val) if not np.isnan(val) else None
                    record[db_col] = val
            records.append(record)
        
        # Batch insert
        batch_size = 500
        inserted = 0
        
        for i in range(0, len(records), batch_size):
            batch = records[i:i+batch_size]
            try:
                supabase.table('envanter_veri').insert(batch).execute()
                inserted += len(batch)
            except Exception as e:
                st.warning(f"Batch {i//batch_size + 1} hatası: {str(e)[:100]}")
        
        return inserted, len(skipped_inventories), new_inventories
        
    except Exception as e:
        return 0, 0, f"Hata: {str(e)}"


def get_available_periods_from_supabase():
    """Mevcut dönemleri al"""
    try:
        result = supabase.table('envanter_veri').select('envanter_donemi').execute()
        if result.data:
            periods = list(set([r['envanter_donemi'] for r in result.data if r['envanter_donemi']]))
            return sorted(periods, reverse=True)
    except:
        pass
    return []


def get_available_sms_from_supabase():
    """Mevcut Satış Müdürlerini al"""
    try:
        result = supabase.table('envanter_veri').select('satis_muduru').execute()
        if result.data:
            sms = list(set([r['satis_muduru'] for r in result.data if r['satis_muduru']]))
            return sorted(sms)
    except:
        pass
    return []


def get_data_from_supabase(satis_muduru=None, donemler=None):
    """Supabase'den veri çek ve DataFrame'e çevir - Pagination ile tüm veriyi al"""
    try:
        all_data = []
        batch_size = 1000
        offset = 0
        
        while True:
            # Sorgu oluştur
            query = supabase.table('envanter_veri').select('*')
            
            if satis_muduru:
                query = query.eq('satis_muduru', satis_muduru)
            
            # Dönem filtresi
            if donemler and len(donemler) > 0:
                query = query.in_('envanter_donemi', donemler)
            
            # Pagination
            query = query.range(offset, offset + batch_size - 1)
            
            result = query.execute()
            
            if not result.data or len(result.data) == 0:
                break
            
            all_data.extend(result.data)
            
            # Son batch'te batch_size'dan az veri geldiyse bitir
            if len(result.data) < batch_size:
                break
            
            offset += batch_size
        
        if not all_data:
            return pd.DataFrame()
        
        df = pd.DataFrame(all_data)
        
        # Sütun isimlerini geri çevir
        reverse_mapping = {
            'magaza_kodu': 'Mağaza Kodu',
            'magaza_tanim': 'Mağaza Adı',
            'satis_muduru': 'Satış Müdürü',
            'bolge_sorumlusu': 'Bölge Sorumlusu',
            'depolama_kosulu_grubu': 'Depolama Koşulu Grubu',
            'depolama_kosulu': 'Depolama Koşulu',
            'envanter_donemi': 'Envanter Dönemi',
            'envanter_tarihi': 'Envanter Tarihi',
            'envanter_baslangic_tarihi': 'Envanter Başlangıç Tarihi',
            'urun_grubu_kodu': 'Ürün Grubu Kodu',
            'urun_grubu_tanimi': 'Ürün Grubu',
            'mal_grubu_kodu': 'Mal Grubu Kodu',
            'mal_grubu_tanimi': 'Mal Grubu Tanımı',
            'malzeme_kodu': 'Malzeme Kodu',
            'malzeme_tanimi': 'Malzeme Adı',
            'satis_fiyati': 'Satış Fiyatı',
            'sayim_miktari': 'Sayım Miktarı',
            'sayim_tutari': 'Sayım Tutarı',
            'kaydi_miktar': 'Kaydi Miktar',
            'kaydi_tutar': 'Kaydi Tutar',
            'fark_miktari': 'Fark Miktarı',
            'fark_tutari': 'Fark Tutarı',
            'kismi_envanter_miktari': 'Kısmi Envanter Miktarı',
            'kismi_envanter_tutari': 'Kısmi Envanter Tutarı',
            'fire_miktari': 'Fire Miktarı',
            'fire_tutari': 'Fire Tutarı',
            'onceki_fark_miktari': 'Önceki Fark Miktarı',
            'onceki_fark_tutari': 'Önceki Fark Tutarı',
            'onceki_fire_miktari': 'Önceki Fire Miktarı',
            'onceki_fire_tutari': 'Önceki Fire Tutarı',
            'satis_miktari': 'Satış Miktarı',
            'satis_hasilati': 'Satış Tutarı',
            'iade_miktari': 'İade Miktarı',
            'iade_tutari': 'İade Tutarı',
            'iptal_fisteki_miktar': 'İptal Fişteki Miktar',
            'iptal_fis_tutari': 'İptal Fiş Tutarı',
            'iptal_gp_miktari': 'İptal GP Miktarı',
            'iptal_gp_tutari': 'İptal GP Tutarı',
            'iptal_satir_miktari': 'İptal Satır Miktarı',
            'iptal_satir_tutari': 'İptal Satır Tutarı',
        }
        
        df = df.rename(columns=reverse_mapping)
        
        return df
        
    except Exception as e:
        st.error(f"Supabase hatası: {str(e)}")
        return pd.DataFrame()


# ==================== ANA UYGULAMA ====================

# Çıkış butonu sağ üstte
col_title, col_user = st.columns([4, 1])
with col_title:
    st.title("🔍 Envanter Risk Analizi")
with col_user:
    st.markdown(f"👤 **{st.session_state.user.upper()}**")
    if st.button("🚪 Çıkış", key="logout_btn"):
        st.session_state.user = None
        st.rerun()

# Mobil uyumlu CSS
st.markdown("""
<style>
    .risk-kritik { background-color: #ff4444; color: white; padding: 10px; border-radius: 5px; text-align: center; font-weight: bold; }
    .risk-riskli { background-color: #ff8800; color: white; padding: 10px; border-radius: 5px; text-align: center; font-weight: bold; }
    .risk-dikkat { background-color: #ffcc00; color: black; padding: 10px; border-radius: 5px; text-align: center; font-weight: bold; }
    .risk-temiz { background-color: #00cc66; color: white; padding: 10px; border-radius: 5px; text-align: center; font-weight: bold; }
    
    /* Mobil uyumluluk */
    @media (max-width: 768px) {
        .stMetric { font-size: 0.8rem; }
        .stDataFrame { font-size: 0.7rem; }
        div[data-testid="column"] { padding: 0.25rem !important; }
    }
    
    /* Tablo kaydırma */
    .stDataFrame { overflow-x: auto; }
</style>
""", unsafe_allow_html=True)

# Mod seçimi - SM Özet eklendi
analysis_mode = st.radio("📊 Analiz Modu", ["🏪 Tek Mağaza", "🌍 Bölge Özeti", "👔 SM Özet"], horizontal=True)

# SM Özet modu için dosya yükleme gerekmez
if analysis_mode != "👔 SM Özet":
    # Dosya yükleme - direkt ekranda
    uploaded_file = st.file_uploader("📁 Excel dosyası yükleyin", type=['xlsx', 'xls'])
else:
    uploaded_file = None


def analyze_inventory(df):
    """Veriyi analiz için hazırla"""
    df = df.copy()
    
    col_mapping = {
        'Mağaza Kodu': 'Mağaza Kodu',
        'Mağaza Tanım': 'Mağaza Adı',
        'Malzeme Kodu': 'Malzeme Kodu',
        'Malzeme Tanımı': 'Malzeme Adı',
        'Mal Grubu Tanımı': 'Ürün Grubu',
        'Ürün Grubu Tanımı': 'Ana Grup',
        'Fark Miktarı': 'Fark Miktarı',
        'Fark Tutarı': 'Fark Tutarı',
        'Kısmi Envanter Miktarı': 'Kısmi Envanter Miktarı',
        'Kısmi Envanter Tutarı': 'Kısmi Envanter Tutarı',
        'Önceki Fark Miktarı': 'Önceki Fark Miktarı',
        'Önceki Fark Tutarı': 'Önceki Fark Tutarı',
        'Önceki Fire Miktarı': 'Önceki Fire Miktarı',
        'Önceki Fire Tutarı': 'Önceki Fire Tutarı',
        'İptal Satır Miktarı': 'İptal Satır Miktarı',
        'İptal Satır Tutarı': 'İptal Satır Tutarı',
        'Fire Miktarı': 'Fire Miktarı',
        'Fire Tutarı': 'Fire Tutarı',
        'Satış Miktarı': 'Satış Miktarı',
        'Satış Hasılatı': 'Satış Tutarı',
        'Satış Fiyatı': 'Birim Fiyat',
        'Fark+Fire+Kısmi Envanter Tutarı': 'NET_ENVANTER_ETKİ_TUTARI',
        'Envanter Dönemi': 'Envanter Dönemi',
        'Envanter Tarihi': 'Envanter Tarihi',
    }
    
    for old_col, new_col in col_mapping.items():
        if old_col in df.columns:
            df[new_col] = df[old_col]
    
    numeric_cols = ['Fark Miktarı', 'Fark Tutarı', 'Kısmi Envanter Miktarı', 'Kısmi Envanter Tutarı',
                    'Önceki Fark Miktarı', 'Önceki Fark Tutarı', 'İptal Satır Miktarı', 'İptal Satır Tutarı',
                    'Fire Miktarı', 'Fire Tutarı', 'Satış Miktarı', 'Satış Tutarı', 'Önceki Fire Miktarı', 
                    'Önceki Fire Tutarı', 'Birim Fiyat']
    
    for col in numeric_cols:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    if 'NET_ENVANTER_ETKİ_TUTARI' not in df.columns:
        df['NET_ENVANTER_ETKİ_TUTARI'] = df['Fark Tutarı'] + df['Fire Tutarı'] + df['Kısmi Envanter Tutarı']
    
    df['TOPLAM_MIKTAR'] = df['Fark Miktarı'] + df['Kısmi Envanter Miktarı'] + df['Önceki Fark Miktarı']
    
    return df


def is_balanced(row):
    """Dengelenmiş mi? Fark + Kısmi + Önceki = 0"""
    toplam = row['Fark Miktarı'] + row['Kısmi Envanter Miktarı'] + row['Önceki Fark Miktarı']
    return abs(toplam) <= 0.01


def get_first_two_words(text):
    """İlk 2 kelimeyi al"""
    if pd.isna(text):
        return ""
    words = str(text).strip().split()
    return " ".join(words[:2]).upper() if len(words) >= 2 else str(text).upper()


def get_last_word(text):
    """Son kelimeyi (marka) al"""
    if pd.isna(text):
        return ""
    words = str(text).strip().split()
    return words[-1].upper() if words else ""


def extract_quantity(text):
    """Gramaj/ML çıkar: '750 ML' → 750, 'ML'"""
    import re
    if pd.isna(text):
        return None, None
    
    text = str(text).upper()
    
    # Patterns: 750ML, 750 ML, 1.5L, 1,5 LT, 220G, 220 G, 1KG
    patterns = [
        r'(\d+[.,]?\d*)\s*(ML|LT|L|G|GR|KG|MG)\b',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            value = float(match.group(1).replace(',', '.'))
            unit = match.group(2)
            
            # Normalize units to base (ML, G)
            if unit in ['LT', 'L']:
                value = value * 1000  # to ML
                unit = 'ML'
            elif unit == 'KG':
                value = value * 1000  # to G
                unit = 'G'
            elif unit == 'GR':
                unit = 'G'
            
            return value, unit
    
    return None, None


def is_quantity_similar(qty1, unit1, qty2, unit2, tolerance=0.30):
    """Gramaj benzer mi? Aynı boyut kategorisinde mi?"""
    if qty1 is None or qty2 is None:
        return True  # Gramaj bulunamadıysa benzer say
    
    if unit1 != unit2:
        return False  # Farklı birim (ML vs G) benzer değil
    
    if qty1 == 0 or qty2 == 0:
        return True
    
    # Oran kontrolü: max 3x fark olabilir
    ratio = max(qty1, qty2) / min(qty1, qty2)
    if ratio > 3:
        return False  # 3 kattan fazla fark varsa benzer değil
    
    # Boyut kategorileri
    def get_size_category(qty, unit):
        if unit == 'ML':
            if qty <= 400: return 'S'      # Küçük: 0-400ml
            elif qty <= 1000: return 'M'   # Orta: 400-1000ml
            else: return 'L'               # Büyük: 1000ml+
        elif unit == 'G':
            if qty <= 100: return 'S'      # Küçük: 0-100g
            elif qty <= 400: return 'M'    # Orta: 100-400g
            else: return 'L'               # Büyük: 400g+
        return 'M'
    
    cat1 = get_size_category(qty1, unit1)
    cat2 = get_size_category(qty2, unit2)
    
    # Sadece aynı kategorideyse benzer
    return cat1 == cat2


def detect_internal_theft(df):
    """
    İÇ HIRSIZLIK TESPİTİ:
    - Satış Fiyatı >= 100 TL
    - Dengelenmemiş (Fark + Kısmi + Önceki ≠ 0)
    - |Toplam| ≈ İptal Satır, fark büyüdükçe risk AZALIR
    """
    results = []
    
    for idx, row in df.iterrows():
        # Dengelenmiş ise atla
        if is_balanced(row):
            continue
        
        satis_fiyati = row.get('Birim Fiyat', 0) or 0
        if satis_fiyati < 100:
            continue
        
        fark = row['Fark Miktarı']
        kismi = row['Kısmi Envanter Miktarı']
        onceki = row['Önceki Fark Miktarı']
        iptal = row['İptal Satır Miktarı']
        
        toplam = fark + kismi + onceki
        
        if toplam >= 0 or iptal <= 0:
            continue
        
        fark_mutlak = abs(abs(toplam) - iptal)
        
        if fark_mutlak == 0:
            risk = "ÇOK YÜKSEK"
            esitlik = "TAM EŞİT"
        elif fark_mutlak <= 2:
            risk = "YÜKSEK"
            esitlik = "YAKIN (±2)"
        elif fark_mutlak <= 5:
            risk = "ORTA"
            esitlik = "YAKIN (±5)"
        elif fark_mutlak <= 10:
            risk = "DÜŞÜK-ORTA"
            esitlik = f"FARK: {fark_mutlak}"
        else:
            continue
        
        results.append({
            'Malzeme Kodu': row.get('Malzeme Kodu', ''),
            'Malzeme Adı': row.get('Malzeme Adı', ''),
            'Ürün Grubu': row.get('Ürün Grubu', ''),
            'Satış Fiyatı': satis_fiyati,
            'Fark Miktarı': fark,
            'Kısmi Env.': kismi,
            'Önceki Fark': onceki,
            'TOPLAM': toplam,
            'İptal Satır': iptal,
            'Fark': fark_mutlak,
            'Durum': esitlik,
            'Fark Tutarı (TL)': row['Fark Tutarı'],
            'Risk': risk
        })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        risk_order = {'ÇOK YÜKSEK': 0, 'YÜKSEK': 1, 'ORTA': 2, 'DÜŞÜK-ORTA': 3}
        result_df['_risk_sort'] = result_df['Risk'].map(risk_order)
        result_df = result_df.sort_values(['_risk_sort', 'Fark Tutarı (TL)'], ascending=[True, True])
        result_df = result_df.drop('_risk_sort', axis=1)
    
    return result_df


def detect_chronic_products(df):
    """Kronik açık - her iki dönemde de Fark < 0"""
    results = []
    
    for idx, row in df.iterrows():
        if is_balanced(row):
            continue
        
        if row['Önceki Fark Miktarı'] < 0 and row['Fark Miktarı'] < 0:
            results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Ürün Grubu', ''),
                'Bu Dönem Fark': row['Fark Miktarı'],
                'Bu Dönem Tutar': row['Fark Tutarı'],
                'Önceki Fark': row['Önceki Fark Miktarı'],
                'Önceki Tutar': row['Önceki Fark Tutarı'],
                'Toplam Tutar': row['Fark Tutarı'] + row['Önceki Fark Tutarı']
            })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Bu Dönem Tutar', ascending=True)
    
    return result_df


def detect_chronic_fire(df):
    """Kronik Fire - her iki dönemde de fire var VE dengelenmemiş"""
    results = []
    
    for idx, row in df.iterrows():
        onceki_fire = row.get('Önceki Fire Miktarı', 0) or 0
        bu_fire = row['Fire Miktarı']
        
        # Her iki dönemde de fire varsa
        if onceki_fire != 0 and bu_fire != 0:
            # Önceki Fark + Fark = 0 ise dengelenmiş, kronik değil
            onceki_fark = row.get('Önceki Fark Miktarı', 0) or 0
            bu_fark = row['Fark Miktarı']
            
            if abs(onceki_fark + bu_fark) <= 0.01:
                continue  # Dengelenmiş, kronik fire değil
            
            results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Ürün Grubu', ''),
                'Bu Dönem Fire': bu_fire,
                'Bu Dönem Fire Tutarı': row['Fire Tutarı'],
                'Önceki Fire': onceki_fire,
                'Önceki Fire Tutarı': row.get('Önceki Fire Tutarı', 0),
                'Toplam Fire Tutarı': row['Fire Tutarı'] + row.get('Önceki Fire Tutarı', 0)
            })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Bu Dönem Fire Tutarı', ascending=True)
    
    return result_df


def detect_fire_manipulation(df):
    """Fire manipülasyonu: Fire var AMA Fark+Kısmi > 0 VE dengelenmemiş"""
    results = []
    
    for idx, row in df.iterrows():
        fark = row['Fark Miktarı']
        kismi = row['Kısmi Envanter Miktarı']
        onceki_fark = row.get('Önceki Fark Miktarı', 0) or 0
        fire = row['Fire Miktarı']
        
        fark_kismi = fark + kismi
        
        # Önceki Fark + Fark = 0 ise dengelenmiş, manipülasyon değil
        if abs(onceki_fark + fark) <= 0.01:
            continue
        
        if fire < 0 and fark_kismi > 0:
            results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Ürün Grubu': row.get('Ürün Grubu', ''),
                'Fark Miktarı': fark,
                'Kısmi Env.': kismi,
                'Önceki Fark': onceki_fark,
                'Fark + Kısmi': fark_kismi,
                'Fire Miktarı': fire,
                'Fire Tutarı': row['Fire Tutarı'],
                'Sonuç': 'FAZLA FİRE GİRİLMİŞ'
            })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Fire Tutarı', ascending=True)
    
    return result_df


def detect_cigarette_shortage(df):
    """
    Sigara açığı - Tüm sigaraların TOPLAM (Fark + Kısmi + Önceki) değerine bakılır
    Eğer toplam < 0 ise sigara açığı var demektir
    """
    sigara_keywords = ['sigara', 'sıgara', 'cigarette', 'tütün', 'makaron']
    
    # Sigara ürünlerini filtrele - tüm olası sütunları kontrol et
    def is_sigara(row):
        check_cols = ['Ürün Grubu', 'Ana Grup', 'Mal Grubu', 'Mal Grubu Tanımı', 'Malzeme Adı']
        for col in check_cols:
            val = str(row.get(col, '')).lower()
            for kw in sigara_keywords:
                if kw in val:
                    return True
        return False
    
    sigara_mask = df.apply(is_sigara, axis=1)
    
    sigara_df = df[sigara_mask].copy()
    
    if len(sigara_df) == 0:
        return pd.DataFrame()
    
    # Tüm sigaraların toplamını hesapla
    toplam_fark = sigara_df['Fark Miktarı'].fillna(0).sum()
    toplam_kismi = sigara_df['Kısmi Envanter Miktarı'].fillna(0).sum()
    toplam_onceki = sigara_df['Önceki Fark Miktarı'].fillna(0).sum()
    net_toplam = toplam_fark + toplam_kismi + toplam_onceki
    
    # Eğer net toplam < 0 ise açık var
    if net_toplam >= 0:
        return pd.DataFrame()
    
    # Açık varsa, detay göster
    results = []
    for idx, row in sigara_df.iterrows():
        fark = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
        kismi = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
        onceki = row['Önceki Fark Miktarı'] if pd.notna(row['Önceki Fark Miktarı']) else 0
        urun_toplam = fark + kismi + onceki
        
        # Sadece 0 olmayan kayıtları göster
        if fark != 0 or kismi != 0 or onceki != 0:
            results.append({
                'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                'Malzeme Adı': row.get('Malzeme Adı', ''),
                'Fark': fark,
                'Kısmi': kismi,
                'Önceki': onceki,
                'Ürün Toplam': urun_toplam,
                'Risk': 'SİGARA'
            })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Ürün Toplam', ascending=True)
        # En sona toplam satırı ekle
        toplam_row = pd.DataFrame([{
            'Malzeme Kodu': '*** TOPLAM ***',
            'Malzeme Adı': f'SİGARA AÇIĞI: {abs(net_toplam):.0f} adet',
            'Fark': toplam_fark,
            'Kısmi': toplam_kismi,
            'Önceki': toplam_onceki,
            'Ürün Toplam': net_toplam,
            'Risk': '⚠️ AÇIK VAR'
        }])
        result_df = pd.concat([result_df, toplam_row], ignore_index=True)
    
    return result_df


def find_product_families(df):
    """
    Benzer ürün ailesi analizi
    Kural: İlk 2 kelime + Son kelime (marka) + Mal Grubu + Gramaj (±%30) aynıysa = AİLE
    """
    df_copy = df.copy()
    df_copy['İlk2Kelime'] = df_copy['Malzeme Adı'].apply(get_first_two_words)
    df_copy['Marka'] = df_copy['Malzeme Adı'].apply(get_last_word)
    df_copy['Gramaj'] = df_copy['Malzeme Adı'].apply(lambda x: extract_quantity(x)[0])
    df_copy['GramajBirim'] = df_copy['Malzeme Adı'].apply(lambda x: extract_quantity(x)[1])
    
    families = []
    processed_indices = set()
    
    # Her ürün için potansiyel aile bul
    for idx, row in df_copy.iterrows():
        if idx in processed_indices:
            continue
        
        ilk2 = row['İlk2Kelime']
        marka = row['Marka']
        urun_grubu = row['Ürün Grubu']
        gramaj = row['Gramaj']
        birim = row['GramajBirim']
        
        if not ilk2 or not marka:
            continue
        
        # Aynı grup içinde benzer ürünleri bul
        family_mask = (
            (df_copy['İlk2Kelime'] == ilk2) & 
            (df_copy['Marka'] == marka) & 
            (df_copy['Ürün Grubu'] == urun_grubu)
        )
        
        potential_family = df_copy[family_mask]
        
        if len(potential_family) <= 1:
            continue
        
        # Gramaj kontrolü - benzer gramajlı olanları filtrele
        family_members = []
        for fam_idx, fam_row in potential_family.iterrows():
            if is_quantity_similar(gramaj, birim, fam_row['Gramaj'], fam_row['GramajBirim']):
                family_members.append(fam_idx)
                processed_indices.add(fam_idx)
        
        if len(family_members) <= 1:
            continue
        
        family_df = df_copy.loc[family_members]
        
        toplam_fark = family_df['Fark Miktarı'].sum()
        toplam_kismi = family_df['Kısmi Envanter Miktarı'].sum()
        toplam_onceki = family_df['Önceki Fark Miktarı'].sum()
        aile_toplami = toplam_fark + toplam_kismi + toplam_onceki
        
        if family_df['Fark Miktarı'].abs().sum() > 0:
            if abs(aile_toplami) <= 2:
                sonuc = "KOD KARIŞIKLIĞI - HIRSIZLIK DEĞİL"
                risk = "DÜŞÜK"
            elif aile_toplami < -2:
                sonuc = "AİLEDE NET AÇIK VAR"
                risk = "ORTA"
            else:
                sonuc = "AİLEDE FAZLA VAR"
                risk = "DÜŞÜK"
            
            urunler = family_df['Malzeme Adı'].tolist()
            farklar = family_df['Fark Miktarı'].tolist()
            
            families.append({
                'Mal Grubu': urun_grubu,
                'İlk 2 Kelime': ilk2,
                'Marka': marka,
                'Ürün Sayısı': len(family_members),
                'Toplam Fark': toplam_fark,
                'Toplam Kısmi': toplam_kismi,
                'Toplam Önceki': toplam_onceki,
                'AİLE TOPLAMI': aile_toplami,
                'Sonuç': sonuc,
                'Risk': risk,
                'Ürünler': ' | '.join([f"{u[:25]}({f})" for u, f in zip(urunler[:5], farklar[:5])])
            })
    
    result_df = pd.DataFrame(families)
    if len(result_df) > 0:
        result_df = result_df.sort_values('AİLE TOPLAMI', ascending=True)
    
    return result_df


def detect_external_theft(df):
    """Dış hırsızlık - açık var ama fire/iptal yok"""
    results = []
    
    for idx, row in df.iterrows():
        if is_balanced(row):
            continue
        
        if row['Fark Miktarı'] < 0 and row['Fire Miktarı'] == 0 and row['İptal Satır Miktarı'] == 0:
            if abs(row['Fark Tutarı']) > 50:
                results.append({
                    'Malzeme Kodu': row.get('Malzeme Kodu', ''),
                    'Malzeme Adı': row.get('Malzeme Adı', ''),
                    'Ürün Grubu': row.get('Ürün Grubu', ''),
                    'Fark Miktarı': row['Fark Miktarı'],
                    'Fark Tutarı': row['Fark Tutarı'],
                    'Önceki Fark': row['Önceki Fark Miktarı'],
                    'Risk': 'DIŞ HIRSIZLIK / SAYIM HATASI'
                })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Fark Tutarı', ascending=True)
    
    return result_df


def check_kasa_activity_products(df, kasa_kodlari):
    """
    10 TL Ürünleri Kontrolü
    Fiyat değişikliği olan ürünlerde manipülasyon riski
    Toplam adet ve tutar etkisini hesapla
    FORMÜL: Fark + Kısmi (Önceki dahil değil)
    """
    results = []
    
    toplam_adet = 0
    toplam_tutar = 0
    eslesen_urun = 0
    
    for idx, row in df.iterrows():
        # Kod eşleştirme - hem string hem int formatını dene
        kod_raw = row.get('Malzeme Kodu', '')
        kod_str = str(kod_raw).replace('.0', '').strip()  # Float'tan gelen .0'ı kaldır
        
        if kod_str in kasa_kodlari:
            eslesen_urun += 1
            fark = row['Fark Miktarı'] if pd.notna(row['Fark Miktarı']) else 0
            kismi = row['Kısmi Envanter Miktarı'] if pd.notna(row['Kısmi Envanter Miktarı']) else 0
            toplam = fark + kismi  # Önceki dahil değil!
            
            # Tutar hesabı - Fark + Kısmi tutarları
            fark_tutari = row.get('Fark Tutarı', 0) or 0
            kismi_tutari = row.get('Kısmi Envanter Tutarı', 0) or 0
            urun_toplam_tutar = fark_tutari + kismi_tutari  # Önceki dahil değil!
            
            toplam_adet += toplam
            toplam_tutar += urun_toplam_tutar
            
            if toplam != 0:  # Sadece sıfır olmayanları göster
                if toplam > 0:
                    durum = "FAZLA (+)"
                else:
                    durum = "AÇIK (-)"
                
                results.append({
                    'Malzeme Kodu': kod_str,
                    'Malzeme Adı': row.get('Malzeme Adı', ''),
                    'Fark': fark,
                    'Kısmi': kismi,
                    'TOPLAM': toplam,
                    'Tutar': urun_toplam_tutar,
                    'Durum': durum
                })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        # Önce fazla (+) olanlar, sonra açık (-) olanlar
        result_df['_sort'] = result_df['TOPLAM'].apply(lambda x: 0 if x > 0 else 1)
        result_df = result_df.sort_values(['_sort', 'TOPLAM'], ascending=[True, False])
        result_df = result_df.drop('_sort', axis=1)
    
    # Özet bilgileri de döndür
    summary = {
        'toplam_urun': eslesen_urun,
        'sorunlu_urun': len(results),
        'toplam_adet': toplam_adet,
        'toplam_tutar': toplam_tutar
    }
    
    return result_df, summary


# 10 TL Ürünleri Ürün Kodları (209 adet)
# Bu ürünlerde fiyat değişikliği olduğu için manipülasyon riski var
KASA_AKTIVITESI_KODLARI = {
    '25006448', '12002256', '12002046', '22001972', '12003295', '22002759', '22002500', '11002886', '22002215', '22002214',
    '22002259', '22002349', '16002163', '22002717', '16001587', '13001073', '30000944', '18002488', '17003609', '22002296',
    '22002652', '24004136', '24004137', '12003073', '22002328', '24005228', '24006215', '24005232', '24005231', '24006214',
    '24006212', '16002332', '16002342', '23001397', '16002310', '24001063', '24004020', '13002613', '13002317', '13002506',
    '16002285', '16002219', '16002286', '16002218', '13000258', '13000257', '13000256', '13000260', '13002533', '22002611',
    '22002579', '13002559', '13000187', '13002904', '13000189', '13000190', '13002908', '13001872', '13001874', '30000838',
    '30000926', '22002605', '22002604', '22002603', '12003241', '16002194', '16001734', '25005580', '25000237', '25000049',
    '16002099', '23001367', '23001510', '23001177', '23001403', '23001278', '22002732', '22002576', '22002577', '25006483',
    '23001240', '16002317', '30000958', '30000956', '24005155', '24005154', '24005156', '24005157', '24005153', '22000280',
    '22002773', '22002774', '22002501', '22002225', '22000397', '22001395', '22000396', '16001859', '18002956', '17003542',
    '16002338', '16002339', '16002341', '16002009', '16000856', '22002715', '16002235', '24006067', '24006069', '24006068',
    '24006066', '22002686', '22002687', '22002688', '16002220', '24005291', '24005290', '24006078', '24006084', '24005288',
    '24006082', '24006079', '24005289', '24006085', '22002763', '22002762', '22001032', '18003049', '24006126', '24004420',
    '24005183', '24005649', '24005650', '14002481', '13002315', '22001229', '13002478', '30000880', '24005798', '24005796',
    '24005799', '24005797', '24005795', '24006159', '24003492', '24006171', '24006170', '24006174', '24006172', '24006173',
    '22002640', '22002553', '22002764', '22002223', '22002679', '22002221', '22002224', '22002572', '27002662', '24005441',
    '24005897', '24005898', '24005900', '24006081', '24006080', '16002087', '22002282', '22002283', '24005893', '24005894',
    '23001198', '23001439', '23001195', '23001199', '23000843', '23000034', '23001445', '23001444', '23001443', '23001522',
    '24004381', '24005184', '23001534', '23001533', '18001591', '27002676', '27002677', '16001956', '24003287', '24000005',
    '24002194', '24002192', '24002764', '24003872', '16001983', '18002969', '27001340', '27001148', '27001563', '24004354',
    '24004196', '24004115', '14002424', '24003641', '24004972', '13001481', '24003327', '24000004', '23000122',
}


def load_kasa_activity_codes():
    """Kasa aktivitesi ürün kodlarını döndür"""
    return KASA_AKTIVITESI_KODLARI


def generate_executive_summary(df, kasa_activity_df=None, kasa_summary=None):
    """Yönetici özeti - mal grubu bazlı yorumlar"""
    comments = []
    
    # Önce toplam tutarı hesapla (Fark + Kısmi + Önceki)
    df_copy = df.copy()
    df_copy['Kısmi Envanter Tutarı'] = df_copy.get('Kısmi Envanter Tutarı', 0).fillna(0)
    df_copy['Önceki Fark Tutarı'] = df_copy.get('Önceki Fark Tutarı', 0).fillna(0)
    df_copy['Toplam Tutar'] = df_copy['Fark Tutarı'] + df_copy['Kısmi Envanter Tutarı'] + df_copy['Önceki Fark Tutarı']
    
    # Mal grubu bazlı analiz
    group_stats = df_copy.groupby('Ürün Grubu').agg({
        'Toplam Tutar': 'sum',
        'Fire Tutarı': 'sum',
        'Satış Tutarı': 'sum',
        'Fark Miktarı': lambda x: (x < 0).sum()
    }).reset_index()
    
    group_stats.columns = ['Ürün Grubu', 'Toplam Fark', 'Toplam Fire', 'Toplam Satış', 'Açık Ürün Sayısı']
    group_stats['Açık Oranı'] = abs(group_stats['Toplam Fark']) / group_stats['Toplam Satış'].replace(0, 1) * 100
    
    # En yüksek açık
    top_acik = group_stats.nsmallest(3, 'Toplam Fark')
    for _, row in top_acik.iterrows():
        if row['Toplam Fark'] < -500:
            comments.append(f"⚠️ {row['Ürün Grubu']}: {row['Toplam Fark']:,.0f} TL açık ({row['Açık Ürün Sayısı']} ürün)")
    
    # En yüksek fire
    top_fire = group_stats.nsmallest(3, 'Toplam Fire')
    for _, row in top_fire.iterrows():
        if row['Toplam Fire'] < -500:
            comments.append(f"🔥 {row['Ürün Grubu']}: {row['Toplam Fire']:,.0f} TL fire")
    
    # 10 TL ürünleri yorumu - TOPLAM ADET VE TUTAR
    if kasa_summary is not None:
        toplam_adet = kasa_summary.get('toplam_adet', 0)
        toplam_tutar = kasa_summary.get('toplam_tutar', 0)
        
        if toplam_adet > 0:
            comments.append(f"💰 10 TL ÜRÜNLERİ: NET +{toplam_adet:.0f} adet / {toplam_tutar:,.0f} TL FAZLA")
            comments.append(f"   ⚠️ Bu fazlalık gerçek envanter açığını gizliyor olabilir!")
        elif toplam_adet < 0:
            comments.append(f"💰 10 TL ÜRÜNLERİ: NET {toplam_adet:.0f} adet / {toplam_tutar:,.0f} TL AÇIK")
    
    return comments, group_stats


def analyze_region(df, kasa_kodlari):
    """Bölge geneli analiz - tüm mağazaları karşılaştır"""
    
    magazalar = df['Mağaza Kodu'].dropna().unique().tolist()
    results = []
    
    for mag in magazalar:
        df_mag = df[df['Mağaza Kodu'] == mag].copy()
        
        if len(df_mag) == 0:
            continue
        
        # Mağaza adı ve BS
        mag_adi = df_mag['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df_mag.columns else ''
        bs = df_mag['Bölge Sorumlusu'].iloc[0] if 'Bölge Sorumlusu' in df_mag.columns else ''
        
        # Gün hesabı
        gun_sayisi = 1
        try:
            if 'Envanter Tarihi' in df_mag.columns and 'Envanter Başlangıç Tarihi' in df_mag.columns:
                env_tarihi = pd.to_datetime(df_mag['Envanter Tarihi'].iloc[0])
                env_baslangic = pd.to_datetime(df_mag['Envanter Başlangıç Tarihi'].iloc[0])
                gun_sayisi = (env_tarihi - env_baslangic).days
                if gun_sayisi <= 0:
                    gun_sayisi = 1
        except:
            gun_sayisi = 1
        
        # Temel metrikler
        toplam_satis = df_mag['Satış Tutarı'].sum()
        
        # Fark = Fark Tutarı + Kısmi Envanter Tutarı
        fark_tutari = df_mag['Fark Tutarı'].fillna(0).sum()
        kismi_tutari = df_mag['Kısmi Envanter Tutarı'].fillna(0).sum()
        fark = fark_tutari + kismi_tutari
        
        # Fire = Fire Tutarı
        fire = df_mag['Fire Tutarı'].fillna(0).sum()
        
        # Toplam Açık = Fark + Fire
        toplam_acik = fark + fire
        
        # Oranlar
        fark_oran = abs(fark) / toplam_satis * 100 if toplam_satis > 0 else 0
        fire_oran = abs(fire) / toplam_satis * 100 if toplam_satis > 0 else 0
        toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
        
        # Günlük hesaplar
        gunluk_fark = fark / gun_sayisi
        gunluk_fire = fire / gun_sayisi
        
        # Risk analizleri
        internal_df = detect_internal_theft(df_mag)
        chronic_df = detect_chronic_products(df_mag)
        chronic_fire_df = detect_chronic_fire(df_mag)
        cigarette_df = detect_cigarette_shortage(df_mag)
        fire_manip_df = detect_fire_manipulation(df_mag)
        kasa_df, kasa_sum = check_kasa_activity_products(df_mag, kasa_kodlari)
        
        # Risk puanı hesapla (ağırlıklı)
        risk_puan = 0
        risk_nedenler = []
        
        # Toplam oran bazlı risk
        if toplam_oran > 2:
            risk_puan += 40
            risk_nedenler.append(f"Toplam %{toplam_oran:.1f}")
        elif toplam_oran > 1.5:
            risk_puan += 25
            risk_nedenler.append(f"Toplam %{toplam_oran:.1f}")
        elif toplam_oran > 1:
            risk_puan += 15
        
        # İç hırsızlık
        if len(internal_df) > 50:
            risk_puan += 30
            risk_nedenler.append(f"İç hırs. {len(internal_df)}")
        elif len(internal_df) > 30:
            risk_puan += 20
            risk_nedenler.append(f"İç hırs. {len(internal_df)}")
        elif len(internal_df) > 15:
            risk_puan += 10
        
        # Sigara açığı (kritik!) - Toplam bazlı
        # cigarette_df boş değilse, içindeki son satırda toplam var
        sigara_acik = 0
        if len(cigarette_df) > 0 and 'Ürün Toplam' in cigarette_df.columns:
            # Son satırdaki Net Toplam değerini al (negatif)
            son_satir = cigarette_df.iloc[-1]
            if son_satir['Malzeme Kodu'] == '*** TOPLAM ***':
                sigara_acik = abs(son_satir['Ürün Toplam'])
        
        if sigara_acik > 5:
            risk_puan += 35
            risk_nedenler.append(f"🚬 SİGARA {sigara_acik:.0f}")
        elif sigara_acik > 0:
            risk_puan += 20
            risk_nedenler.append(f"🚬 Sigara {sigara_acik:.0f}")
        
        # Kronik açık
        if len(chronic_df) > 100:
            risk_puan += 15
            risk_nedenler.append(f"Kronik {len(chronic_df)}")
        elif len(chronic_df) > 50:
            risk_puan += 10
        
        # Fire manipülasyonu
        if len(fire_manip_df) > 10:
            risk_puan += 20
            risk_nedenler.append(f"Fire man. {len(fire_manip_df)}")
        elif len(fire_manip_df) > 5:
            risk_puan += 10
        
        # 10 TL ürünleri (fazla = şüpheli)
        if kasa_sum['toplam_adet'] > 20:
            risk_puan += 15
            risk_nedenler.append(f"10TL +{kasa_sum['toplam_adet']:.0f}")
        elif kasa_sum['toplam_adet'] > 10:
            risk_puan += 10
        
        # Risk puanını 100 ile sınırla
        risk_puan = min(risk_puan, 100)
        
        # Risk seviyesi belirleme
        if risk_puan >= 60:
            risk_seviye = "🔴 KRİTİK"
        elif risk_puan >= 40:
            risk_seviye = "🟠 RİSKLİ"
        elif risk_puan >= 20:
            risk_seviye = "🟡 DİKKAT"
        else:
            risk_seviye = "🟢 TEMİZ"
        
        results.append({
            'Mağaza Kodu': mag,
            'Mağaza Adı': mag_adi,
            'BS': bs,
            'Satış': toplam_satis,
            'Fark': fark,
            'Fire': fire,
            'Toplam Açık': toplam_acik,
            'Fark %': fark_oran,
            'Fire %': fire_oran,
            'Toplam %': toplam_oran,
            'Gün': gun_sayisi,
            'Günlük Fark': gunluk_fark,
            'Günlük Fire': gunluk_fire,
            'İç Hırs.': len(internal_df),
            'Kr.Açık': len(chronic_df),
            'Kr.Fire': len(chronic_fire_df),
            'Sigara': sigara_acik,
            'Fire Man.': len(fire_manip_df),
            '10TL Adet': kasa_sum['toplam_adet'],
            '10TL Tutar': kasa_sum['toplam_tutar'],
            'Risk Puan': risk_puan,
            'Risk': risk_seviye,
            'Risk Nedenleri': " | ".join(risk_nedenler) if risk_nedenler else "-"
        })
    
    result_df = pd.DataFrame(results)
    if len(result_df) > 0:
        result_df = result_df.sort_values('Risk Puan', ascending=False)
    
    return result_df


def create_region_excel_report(region_df, df_all, kasa_kodlari, params):
    """Bölge özet Excel raporu"""
    
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    kritik_fill = PatternFill('solid', fgColor='FF4444')
    riskli_fill = PatternFill('solid', fgColor='FF8800')
    dikkat_fill = PatternFill('solid', fgColor='FFCC00')
    temiz_fill = PatternFill('solid', fgColor='00CC66')
    title_font = Font(bold=True, size=14)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ===== BÖLGE ÖZETİ =====
    ws = wb.active
    ws.title = "BÖLGE ÖZETİ"
    
    ws['A1'] = f"BÖLGE ENVANTER ANALİZİ"
    ws['A1'].font = title_font
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Tarih: {params.get('tarih', '')} | Mağaza Sayısı: {len(region_df)}"
    
    # Bölge toplamları
    ws['A4'] = "BÖLGE TOPLAMI"
    ws['A4'].font = Font(bold=True, size=11)
    
    toplam_satis = region_df['Satış'].sum()
    toplam_fark = region_df['Fark'].sum()
    toplam_fire = region_df['Fire'].sum()
    # Kayıp Oranı = |Fark + Fire| / Satış × 100
    genel_oran = abs(toplam_fark + toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
    
    ws['A5'] = "Toplam Satış"
    ws['B5'] = f"{toplam_satis:,.0f} TL"
    ws['A6'] = "Toplam Fark"
    ws['B6'] = f"{toplam_fark:,.0f} TL"
    ws['A7'] = "Toplam Fire"
    ws['B7'] = f"{toplam_fire:,.0f} TL"
    ws['A8'] = "Genel Kayıp Oranı"
    ws['B8'] = f"%{genel_oran:.2f}"
    
    # Risk dağılımı
    ws['A10'] = "RİSK DAĞILIMI"
    ws['A10'].font = Font(bold=True, size=11)
    
    kritik_sayisi = len(region_df[region_df['Risk'].str.contains('KRİTİK')])
    riskli_sayisi = len(region_df[region_df['Risk'].str.contains('RİSKLİ')])
    dikkat_sayisi = len(region_df[region_df['Risk'].str.contains('DİKKAT')])
    temiz_sayisi = len(region_df[region_df['Risk'].str.contains('TEMİZ')])
    
    ws['A11'] = "🔴 KRİTİK"
    ws['B11'] = kritik_sayisi
    ws['A12'] = "🟠 RİSKLİ"
    ws['B12'] = riskli_sayisi
    ws['A13'] = "🟡 DİKKAT"
    ws['B13'] = dikkat_sayisi
    ws['A14'] = "🟢 TEMİZ"
    ws['B14'] = temiz_sayisi
    
    # Mağaza sıralaması
    ws['A16'] = "MAĞAZA SIRALAMASI (Risk Puanına Göre)"
    ws['A16'].font = Font(bold=True, size=11)
    
    headers = ['Mağaza', 'Adı', 'Satış', 'Fark', 'Toplam %', 'İç Hırs.', 'Sigara', 'Kr.Açık', 'Risk', 'Neden']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row_idx, (_, row) in enumerate(region_df.iterrows(), start=18):
        ws.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = border
        ws.cell(row=row_idx, column=2, value=row['Mağaza Adı'][:25]).border = border
        ws.cell(row=row_idx, column=3, value=f"{row['Satış']:,.0f}").border = border
        ws.cell(row=row_idx, column=4, value=f"{row['Fark']:,.0f}").border = border
        ws.cell(row=row_idx, column=5, value=f"%{row['Toplam %']:.1f}").border = border
        ws.cell(row=row_idx, column=6, value=row['İç Hırs.']).border = border
        ws.cell(row=row_idx, column=7, value=row['Sigara']).border = border
        ws.cell(row=row_idx, column=8, value=row['Kr.Açık']).border = border
        
        risk_cell = ws.cell(row=row_idx, column=9, value=row['Risk'])
        risk_cell.border = border
        if 'KRİTİK' in row['Risk']:
            risk_cell.fill = kritik_fill
            risk_cell.font = Font(bold=True, color='FFFFFF')
        elif 'RİSKLİ' in row['Risk']:
            risk_cell.fill = riskli_fill
            risk_cell.font = Font(bold=True, color='FFFFFF')
        elif 'DİKKAT' in row['Risk']:
            risk_cell.fill = dikkat_fill
            risk_cell.font = Font(bold=True)
        else:
            risk_cell.fill = temiz_fill
            risk_cell.font = Font(bold=True, color='FFFFFF')
        
        ws.cell(row=row_idx, column=10, value=row['Risk Nedenleri']).border = border
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 35
    
    # ===== DETAY SHEET =====
    ws2 = wb.create_sheet("DETAY")
    
    detail_headers = ['Mağaza Kodu', 'Mağaza Adı', 'Satış', 'Fark', 'Fire', 'Toplam %', 
                      'İç Hırs.', 'Kr.Açık', 'Kr.Fire', 'Sigara', 'Fire Man.', 
                      '10TL Adet', '10TL Tutar', 'Risk Puan', 'Risk', 'Risk Nedenleri']
    
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row_idx, (_, row) in enumerate(region_df.iterrows(), start=2):
        ws2.cell(row=row_idx, column=1, value=row['Mağaza Kodu']).border = border
        ws2.cell(row=row_idx, column=2, value=row['Mağaza Adı']).border = border
        ws2.cell(row=row_idx, column=3, value=row['Satış']).border = border
        ws2.cell(row=row_idx, column=4, value=row['Fark']).border = border
        ws2.cell(row=row_idx, column=5, value=row['Fire']).border = border
        ws2.cell(row=row_idx, column=6, value=row['Toplam %']).border = border
        ws2.cell(row=row_idx, column=7, value=row['İç Hırs.']).border = border
        ws2.cell(row=row_idx, column=8, value=row['Kr.Açık']).border = border
        ws2.cell(row=row_idx, column=9, value=row['Kr.Fire']).border = border
        ws2.cell(row=row_idx, column=10, value=row['Sigara']).border = border
        ws2.cell(row=row_idx, column=11, value=row['Fire Man.']).border = border
        ws2.cell(row=row_idx, column=12, value=row['10TL Adet']).border = border
        ws2.cell(row=row_idx, column=13, value=row['10TL Tutar']).border = border
        ws2.cell(row=row_idx, column=14, value=row['Risk Puan']).border = border
        ws2.cell(row=row_idx, column=15, value=row['Risk']).border = border
        ws2.cell(row=row_idx, column=16, value=row['Risk Nedenleri']).border = border
    
    auto_adjust_column_width(ws2)
    
    # Excel çıktısı
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def calculate_store_risk(df, internal_df, chronic_df, cigarette_df):
    """Mağaza risk seviyesi"""
    toplam_satis = df['Satış Tutarı'].sum()
    fark_tutari = df['Fark Tutarı'].fillna(0).sum()
    fire_tutari = df['Fire Tutarı'].fillna(0).sum()
    kismi_tutari = df['Kısmi Envanter Tutarı'].fillna(0).sum()
    
    # Kayıp Oranı = |Fark + Fire + Kısmi| / Satış × 100
    kayip = fark_tutari + fire_tutari + kismi_tutari
    kayip_orani = abs(kayip) / toplam_satis * 100 if toplam_satis > 0 else 0
    ic_hirsizlik = len(internal_df)
    
    # Sigara açığı - toplam bazlı
    sigara_acik = 0
    if len(cigarette_df) > 0 and 'Ürün Toplam' in cigarette_df.columns:
        son_satir = cigarette_df.iloc[-1]
        if son_satir['Malzeme Kodu'] == '*** TOPLAM ***':
            sigara_acik = abs(son_satir['Ürün Toplam'])
    
    if kayip_orani > 2 or ic_hirsizlik > 50 or sigara_acik > 5:
        return "KRİTİK", "risk-kritik"
    elif kayip_orani > 1.5 or ic_hirsizlik > 30 or sigara_acik > 3:
        return "RİSKLİ", "risk-riskli"
    elif kayip_orani > 1 or ic_hirsizlik > 15 or sigara_acik > 0:
        return "DİKKAT", "risk-dikkat"
    else:
        return "TEMİZ", "risk-temiz"


def create_top_20_risky(df, internal_codes, chronic_codes, family_balanced_codes):
    """En riskli 20 ürün"""
    
    # Dengelenmişleri ve aile dengelenmişlerini çıkar
    risky_df = df[
        (df['NET_ENVANTER_ETKİ_TUTARI'] < 0) & 
        (~df.apply(is_balanced, axis=1)) &
        (~df['Malzeme Kodu'].astype(str).isin(family_balanced_codes))
    ].copy()
    
    if len(risky_df) == 0:
        return pd.DataFrame()
    
    def classify(row):
        kod = str(row.get('Malzeme Kodu', ''))
        
        if kod in internal_codes:
            return "İÇ HIRSIZLIK", "Kasa kamera incelemesi"
        elif kod in chronic_codes:
            return "KRONİK AÇIK", "Raf kontrolü, Sayım eğitimi"
        elif row['Fire Miktarı'] < 0:
            return "OPERASYONEL", "Fire kayıt kontrolü"
        else:
            return "DIŞ HIRSIZLIK/SAYIM", "Sayım ve kod kontrolü"
    
    risky_df['Risk Türü'] = risky_df.apply(lambda x: classify(x)[0], axis=1)
    risky_df['Aksiyon'] = risky_df.apply(lambda x: classify(x)[1], axis=1)
    
    risky_df = risky_df.sort_values('NET_ENVANTER_ETKİ_TUTARI', ascending=True).head(20)
    
    result = pd.DataFrame({
        'Sıra': range(1, len(risky_df) + 1),
        'Malzeme Kodu': risky_df['Malzeme Kodu'].values,
        'Malzeme Adı': risky_df['Malzeme Adı'].values,
        'Fark Mik.': risky_df['Fark Miktarı'].values,
        'Kısmi': risky_df['Kısmi Envanter Miktarı'].values,
        'Önceki': risky_df['Önceki Fark Miktarı'].values,
        'TOPLAM': risky_df['TOPLAM_MIKTAR'].values,
        'İptal': risky_df['İptal Satır Miktarı'].values,
        'Fire': risky_df['Fire Miktarı'].values,
        'Fire Tutarı': risky_df['Fire Tutarı'].values,
        'Fark Tutarı': risky_df['Fark Tutarı'].values,
        'Risk Türü': risky_df['Risk Türü'].values,
        'Aksiyon': risky_df['Aksiyon'].values
    })
    
    return result


def auto_adjust_column_width(ws):
    """Excel sütun genişliklerini otomatik ayarla"""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


def create_excel_report(df, internal_df, chronic_df, chronic_fire_df, cigarette_df, 
                       external_df, family_df, fire_manip_df, kasa_activity_df, top20_df, 
                       exec_comments, group_stats, magaza_kodu, magaza_adi, params):
    """Excel raporu - tüm sheet'ler dahil"""
    
    wb = Workbook()
    
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # ===== ÖZET =====
    ws = wb.active
    ws.title = "ÖZET"
    
    ws['A1'] = f"MAĞAZA: {magaza_kodu} - {magaza_adi}"
    ws['A1'].font = title_font
    ws['A2'] = f"Dönem: {params.get('donem', '')} | Tarih: {params.get('tarih', '')}"
    
    ws['A4'] = "GENEL METRIKLER"
    ws['A4'].font = subtitle_font
    
    toplam_satis = df['Satış Tutarı'].sum()
    fark_tutari = df['Fark Tutarı'].fillna(0).sum()
    kismi_tutari = df['Kısmi Envanter Tutarı'].fillna(0).sum()
    fire_tutari = df['Fire Tutarı'].fillna(0).sum()
    
    # Fark = Fark Tutarı + Kısmi
    fark = fark_tutari + kismi_tutari
    # Toplam Açık = Fark + Fire
    toplam_acik = fark + fire_tutari
    
    # Oranlar
    fark_oran = abs(fark) / toplam_satis * 100 if toplam_satis > 0 else 0
    fire_oran = abs(fire_tutari) / toplam_satis * 100 if toplam_satis > 0 else 0
    toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
    
    metrics = [
        ('Toplam Ürün', len(df)),
        ('Açık Veren Ürün', len(df[df['Fark Miktarı'] < 0])),
        ('Toplam Satış', f"{toplam_satis:,.0f} TL"),
        ('Fark (Fark+Kısmi)', f"{fark:,.0f} TL"),
        ('Fire', f"{fire_tutari:,.0f} TL"),
        ('Toplam Açık', f"{toplam_acik:,.0f} TL"),
        ('Fark Oranı', f"%{fark_oran:.2f}"),
        ('Fire Oranı', f"%{fire_oran:.2f}"),
        ('Toplam Oran', f"%{toplam_oran:.2f}"),
    ]
    
    for i, (label, value) in enumerate(metrics, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
    
    ws['A15'] = "RİSK DAĞILIMI"
    ws['A15'].font = subtitle_font
    
    risks = [
        ('İç Hırsızlık (≥100TL)', len(internal_df)),
        ('Kronik Açık', len(chronic_df)),
        ('Kronik Fire', len(chronic_fire_df)),
        ('Sigara Açığı', len(cigarette_df)),
        ('Fire Manipülasyonu', len(fire_manip_df)),
    ]
    
    for i, (label, value) in enumerate(risks, start=16):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        if 'Sigara' in label and value > 0:
            ws[f'B{i}'].fill = PatternFill('solid', fgColor='FF4444')
            ws[f'B{i}'].font = Font(bold=True, color='FFFFFF')
    
    ws['A22'] = "YÖNETİCİ ÖZETİ"
    ws['A22'].font = subtitle_font
    
    for i, comment in enumerate(exec_comments[:10], start=23):
        ws[f'A{i}'] = comment
    
    auto_adjust_column_width(ws)
    
    # ===== EN RİSKLİ 20 =====
    if len(top20_df) > 0:
        ws2 = wb.create_sheet("EN RİSKLİ 20")
        for col, h in enumerate(top20_df.columns, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for r_idx, row in enumerate(top20_df.values, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                cell.alignment = wrap_alignment
        
        auto_adjust_column_width(ws2)
    
    # ===== KRONİK AÇIK =====
    if len(chronic_df) > 0:
        ws3 = wb.create_sheet("KRONİK AÇIK")
        for col, h in enumerate(chronic_df.columns, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(chronic_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws3.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws3)
    
    # ===== KRONİK FİRE =====
    if len(chronic_fire_df) > 0:
        ws4 = wb.create_sheet("KRONİK FİRE")
        for col, h in enumerate(chronic_fire_df.columns, 1):
            cell = ws4.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(chronic_fire_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws4.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws4)
    
    # ===== SİGARA AÇIĞI =====
    ws5 = wb.create_sheet("SİGARA AÇIĞI")
    ws5['A1'] = "⚠️ SİGARA AÇIĞI - YÜKSEK RİSK"
    ws5['A1'].font = Font(bold=True, size=14, color='FF0000')
    
    if len(cigarette_df) > 0:
        for col, h in enumerate(cigarette_df.columns, 1):
            cell = ws5.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor='FF4444')
        
        for r_idx, row in enumerate(cigarette_df.values, 4):
            for c_idx, val in enumerate(row, 1):
                ws5.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws5)
    
    # ===== İÇ HIRSIZLIK =====
    if len(internal_df) > 0:
        ws6 = wb.create_sheet("İÇ HIRSIZLIK")
        ws6['A1'] = "Satış Fiyatı ≥ 100 TL | Fark büyüdükçe risk AZALIR"
        ws6['A1'].font = subtitle_font
        
        for col, h in enumerate(internal_df.columns, 1):
            cell = ws6.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(internal_df.head(100).values, 4):
            for c_idx, val in enumerate(row, 1):
                ws6.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws6)
    
    # ===== AİLE ANALİZİ =====
    if len(family_df) > 0:
        ws7 = wb.create_sheet("AİLE ANALİZİ")
        ws7['A1'] = "Benzer Ürün Ailesi - Kod Karışıklığı Tespiti"
        ws7['A1'].font = subtitle_font
        
        for col, h in enumerate(family_df.columns, 1):
            cell = ws7.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(family_df.head(100).values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws7.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = wrap_alignment
        
        auto_adjust_column_width(ws7)
    
    # ===== FİRE MANİPÜLASYONU =====
    if len(fire_manip_df) > 0:
        ws8 = wb.create_sheet("FİRE MANİPÜLASYONU")
        for col, h in enumerate(fire_manip_df.columns, 1):
            cell = ws8.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(fire_manip_df.head(100).values, 2):
            for c_idx, val in enumerate(row, 1):
                ws8.cell(row=r_idx, column=c_idx, value=val)
        
        auto_adjust_column_width(ws8)
    
    # ===== KASA AKTİVİTESİ =====
    if len(kasa_activity_df) > 0:
        ws9 = wb.create_sheet("KASA AKTİVİTESİ")
        ws9['A1'] = "⚠️ KASA AKTİVİTESİ ÜRÜNLERİ - FAZLA (+) OLANLAR MANİPÜLASYON RİSKİ!"
        ws9['A1'].font = Font(bold=True, size=12, color='FF0000')
        
        for col, h in enumerate(kasa_activity_df.columns, 1):
            cell = ws9.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        for r_idx, row in enumerate(kasa_activity_df.values, 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws9.cell(row=r_idx, column=c_idx, value=val)
                # Fazla olanları kırmızı yap
                if c_idx == 6 and isinstance(val, (int, float)) and val > 0:  # TOPLAM sütunu
                    cell.fill = PatternFill('solid', fgColor='FFCCCC')
        
        auto_adjust_column_width(ws9)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ===== ANA UYGULAMA =====

# SM Özet modu - Supabase'den okur
if analysis_mode == "👔 SM Özet":
    st.subheader("👔 SM Özet - Supabase'den")
    
    # Kullanıcı -> SM eşleştirmesi
    USER_SM_MAPPING = {
        "sm1": "ALİ AKÇAY",
        "sm2": "ŞADAN YURDAKUL",
        "sm3": "VELİ GÖK",
        "sm4": "GİZEM TOSUN",
        "sma": None,  # Asistan - SM seçecek
        "ziya": None,  # GM - tüm SM'leri görebilir
    }
    
    current_user = st.session_state.user
    user_sm = USER_SM_MAPPING.get(current_user)
    is_gm = current_user == "ziya"
    
    # SM ve Dönem seçimi - aynı satırda
    col_sm, col_donem = st.columns([1, 1])
    
    available_sms = get_available_sms_from_supabase()
    available_periods = get_available_periods_from_supabase()
    
    with col_sm:
        if is_gm:
            # GM tüm SM'leri görebilir + TÜMÜ seçeneği
            if available_sms:
                sm_options = ["📊 TÜMÜ (Bölge)"] + available_sms
                selected_sm_option = st.selectbox("👔 Satış Müdürü", sm_options)
                
                if selected_sm_option == "📊 TÜMÜ (Bölge)":
                    selected_sm = None
                    display_sm = "Bölge"
                else:
                    selected_sm = selected_sm_option
                    display_sm = selected_sm
            else:
                st.warning("Henüz veri yüklenmemiş")
                selected_sm = None
                selected_sm_option = None
                display_sm = None
        elif user_sm:
            # SM kendi verilerini görür (sadece kendi ismi gösterilir)
            selected_sm = user_sm
            selected_sm_option = user_sm
            display_sm = user_sm
            st.selectbox("👔 Satış Müdürü", [user_sm], disabled=True)
        else:
            # Asistan veya tanımsız - SM seçebilir
            if available_sms:
                selected_sm = st.selectbox("👔 Satış Müdürü", available_sms)
                selected_sm_option = selected_sm
                display_sm = selected_sm
            else:
                st.warning("Henüz veri yüklenmemiş")
                selected_sm = None
                selected_sm_option = None
                display_sm = None
    
    with col_donem:
        if available_periods:
            selected_periods = st.multiselect("📅 Dönem", available_periods, default=available_periods[:1])
        else:
            selected_periods = []
    
    if selected_sm_option and selected_periods:
        with st.spinner("Veriler yükleniyor..."):
            df_supabase = get_data_from_supabase(satis_muduru=selected_sm, donemler=selected_periods)
        
        if len(df_supabase) == 0:
            st.warning("Seçilen kriterlere uygun veri bulunamadı")
        else:
            st.success(f"✅ {len(df_supabase):,} satır yüklendi")
            
            # Veriyi analyze_inventory'den geçir
            df = analyze_inventory(df_supabase)
            
            # Mağaza bilgisi
            if 'Mağaza Kodu' in df.columns:
                magazalar = df['Mağaza Kodu'].dropna().unique().tolist()
                magaza_isimleri = {}
                for mag in magazalar:
                    isim = df[df['Mağaza Kodu'] == mag]['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df.columns else ''
                    magaza_isimleri[mag] = f"{mag} - {isim}" if isim else str(mag)
            else:
                magazalar = []
                magaza_isimleri = {}
            
            params = {
                'donem': ', '.join(selected_periods),
                'tarih': datetime.now().strftime('%Y-%m-%d'),
            }
            
            # Kasa aktivitesi kodlarını yükle
            kasa_kodlari = load_kasa_activity_codes()
            
            # Bölge Özeti ile aynı analiz
            st.subheader(f"📊 {display_sm} - {len(magazalar)} Mağaza")
            
            with st.spinner("Mağazalar analiz ediliyor..."):
                region_df = analyze_region(df, kasa_kodlari)
            
            if len(region_df) == 0:
                st.warning("Analiz edilecek mağaza bulunamadı!")
            else:
                # Bölge toplamları
                toplam_satis = region_df['Satış'].sum()
                toplam_fark = region_df['Fark'].sum()
                toplam_fire = region_df['Fire'].sum()
                toplam_acik = region_df['Toplam Açık'].sum()
                toplam_gun = region_df['Gün'].sum()
                
                # Oranlar
                fark_oran = abs(toplam_fark) / toplam_satis * 100 if toplam_satis > 0 else 0
                fire_oran = abs(toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
                toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
                gunluk_fark = toplam_fark / toplam_gun if toplam_gun > 0 else 0
                gunluk_fire = toplam_fire / toplam_gun if toplam_gun > 0 else 0
                
                # Risk dağılımı
                kritik_sayisi = len(region_df[region_df['Risk'] == '🔴 KRİTİK'])
                riskli_sayisi = len(region_df[region_df['Risk'] == '🟠 RİSKLİ'])
                dikkat_sayisi = len(region_df[region_df['Risk'] == '🟡 DİKKAT'])
                temiz_sayisi = len(region_df[region_df['Risk'] == '🟢 TEMİZ'])
                
                # Üst metrikler
                st.markdown("### 💰 Özet Metrikler")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("💰 Toplam Satış", f"{toplam_satis/1_000_000:.1f}M TL")
                with col2:
                    st.metric("📉 Fark", f"{toplam_fark/1000:.0f}K TL", f"%{fark_oran:.2f} | Günlük: {gunluk_fark/1000:.1f}K")
                with col3:
                    st.metric("🔥 Fire", f"{toplam_fire/1000:.0f}K TL", f"%{fire_oran:.2f} | Günlük: {gunluk_fire/1000:.1f}K")
                with col4:
                    st.metric("📊 Toplam", f"{toplam_acik/1000:.0f}K TL", f"%{toplam_oran:.2f}")
                
                # Risk dağılımı
                st.markdown("### 📊 Risk Dağılımı")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    if kritik_sayisi > 0:
                        st.markdown(f'<div class="risk-kritik">🔴 KRİTİK: {kritik_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🔴 KRİTİK", kritik_sayisi)
                with col2:
                    if riskli_sayisi > 0:
                        st.markdown(f'<div class="risk-riskli">🟠 RİSKLİ: {riskli_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🟠 RİSKLİ", riskli_sayisi)
                with col3:
                    if dikkat_sayisi > 0:
                        st.markdown(f'<div class="risk-dikkat">🟡 DİKKAT: {dikkat_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🟡 DİKKAT", dikkat_sayisi)
                with col4:
                    st.markdown(f'<div class="risk-temiz">🟢 TEMİZ: {temiz_sayisi}</div>', unsafe_allow_html=True)
                
                # BS Özeti
                st.markdown("### 👔 BS Özeti")
                bs_ozet = region_df.groupby('BS').agg({
                    'Mağaza Kodu': 'count',
                    'Satış': 'sum',
                    'Fark': 'sum',
                    'Fire': 'sum',
                    'Toplam Açık': 'sum',
                    'Risk Puan': 'sum',  # Toplam risk puanı
                    'Sigara': 'sum',
                    'İç Hırs.': 'sum'
                }).reset_index()
                bs_ozet.columns = ['BS', 'Mağaza', 'Satış', 'Fark', 'Fire', 'Toplam', 'Risk Puan', 'Sigara', 'İç Hırs.']
                bs_ozet['Kayıp %'] = abs(bs_ozet['Toplam']) / bs_ozet['Satış'] * 100
                bs_ozet = bs_ozet.sort_values('Risk Puan', ascending=False)  # Risk puanına göre sırala
                
                # BS tablosu - tam rakamlar ve risk puanı ile
                for _, bs_row in bs_ozet.iterrows():
                    col1, col2, col3, col4, col5, col6 = st.columns([2.5, 1.5, 1.5, 1, 1, 1])
                    col1.write(f"**{bs_row['BS']}** ({bs_row['Mağaza']:.0f} mağ.)")
                    col2.write(f"Satış: {bs_row['Satış']/1e6:.1f}M | Fark: {bs_row['Fark']:,.0f}")
                    col3.write(f"Fire: {bs_row['Fire']:,.0f}")
                    col4.write(f"Kayıp: %{bs_row['Kayıp %']:.1f}")
                    col5.write(f"🚬{bs_row['Sigara']:.0f} 🔒{bs_row['İç Hırs.']:.0f}")
                    col6.write(f"**Risk: {bs_row['Risk Puan']:.0f}**")
                
                # Sekmeler - Bölge Özeti ile aynı
                st.markdown("---")
                tabs = st.tabs(["📋 Sıralama", "🔴 Kritik", "🟠 Riskli", "🚬 Sigara", "📊 Detay", "📥 İndir"])
                
                with tabs[0]:
                    st.subheader("📋 Mağaza Sıralaması (Risk Puanına Göre)")
                    
                    # Başlık satırı
                    cols = st.columns([0.4, 0.8, 1.3, 1.2, 0.9, 0.7, 0.9, 0.7, 0.6, 0.6, 0.4, 0.5, 0.8])
                    cols[0].markdown("**📥**")
                    cols[1].markdown("**Kod**")
                    cols[2].markdown("**Mağaza Adı**")
                    cols[3].markdown("**BS**")
                    cols[4].markdown("**Fark**")
                    cols[5].markdown("**Günlük**")
                    cols[6].markdown("**Fire**")
                    cols[7].markdown("**Günlük**")
                    cols[8].markdown("**Kayıp%**")
                    cols[9].markdown("**Fire%**")
                    cols[10].markdown("**🚬**")
                    cols[11].markdown("**Risk**")
                    cols[12].markdown("**Seviye**")
                    
                    st.markdown("---")
                    
                    # Veri satırları
                    for idx, (_, row) in enumerate(region_df.iterrows()):
                        cols = st.columns([0.4, 0.8, 1.3, 1.2, 0.9, 0.7, 0.9, 0.7, 0.6, 0.6, 0.4, 0.5, 0.8])
                        
                        # Mağaza verisini al ve tam rapor oluştur
                        mag_kod = row['Mağaza Kodu']
                        df_mag = df[df['Mağaza Kodu'] == mag_kod].copy()
                        mag_adi = row['Mağaza Adı']
                        
                        # Analizleri yap
                        int_df = detect_internal_theft(df_mag)
                        chr_df = detect_chronic_products(df_mag)
                        chr_fire_df = detect_chronic_fire(df_mag)
                        cig_df = detect_cigarette_shortage(df_mag)
                        ext_df = detect_external_theft(df_mag)
                        fam_df = find_product_families(df_mag)
                        fire_df = detect_fire_manipulation(df_mag)
                        kasa_df, kasa_sum = check_kasa_activity_products(df_mag, kasa_kodlari)
                        
                        int_codes = set(int_df['Malzeme Kodu'].astype(str).tolist()) if len(int_df) > 0 else set()
                        chr_codes = set(chr_df['Malzeme Kodu'].astype(str).tolist()) if len(chr_df) > 0 else set()
                        
                        t20_df = create_top_20_risky(df_mag, int_codes, chr_codes, set())
                        exec_c, grp_s = generate_executive_summary(df_mag, kasa_df, kasa_sum)
                        
                        # Tam rapor oluştur
                        report_data = create_excel_report(
                            df_mag, int_df, chr_df, chr_fire_df, cig_df,
                            ext_df, fam_df, fire_df, kasa_df, t20_df,
                            exec_c, grp_s, mag_kod, mag_adi, params
                        )
                        
                        mag_adi_clean = mag_adi.replace(' ', '_').replace('/', '_')[:30] if mag_adi else ''
                        
                        with cols[0]:
                            st.download_button("📥", data=report_data, 
                                file_name=f"{mag_kod}_{mag_adi_clean}_Risk_Raporu.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"sm_dl_{idx}")
                        cols[1].write(f"{row['Mağaza Kodu']}")
                        cols[2].write(f"{row['Mağaza Adı'][:18] if row['Mağaza Adı'] else '-'}")
                        cols[3].write(f"{row['BS'][:12] if row['BS'] else '-'}")
                        cols[4].write(f"{row['Fark']/1000:.0f}K")
                        cols[5].write(f"{row['Günlük Fark']/1000:.1f}K")
                        cols[6].write(f"{row['Fire']/1000:.0f}K")
                        cols[7].write(f"{row['Günlük Fire']/1000:.1f}K")
                        cols[8].write(f"%{row['Toplam %']:.1f}")
                        cols[9].write(f"%{row['Fire %']:.1f}")
                        cols[10].write(f"{row['Sigara']:.0f}" if row['Sigara'] > 0 else "-")
                        cols[11].write(f"{row['Risk Puan']:.0f}")
                        cols[12].write(row['Risk'])
                
                with tabs[1]:
                    st.subheader("🔴 Kritik Mağazalar")
                    kritik_df = region_df[region_df['Risk'].str.contains('KRİTİK')]
                    if len(kritik_df) > 0:
                        for _, row in kritik_df.iterrows():
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**\n\n"
                                    f"Kayıp: %{row['Toplam %']:.1f} | Fark: {row['Fark']:,.0f} TL\n\n"
                                    f"**Neden:** {row['Risk Nedenleri']}")
                    else:
                        st.success("Kritik mağaza yok! 🎉")
                
                with tabs[2]:
                    st.subheader("🟠 Riskli Mağazalar")
                    riskli_df = region_df[region_df['Risk'].str.contains('RİSKLİ')]
                    if len(riskli_df) > 0:
                        for _, row in riskli_df.iterrows():
                            st.warning(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**\n\n"
                                      f"Kayıp: %{row['Toplam %']:.1f} | Fark: {row['Fark']:,.0f} TL\n\n"
                                      f"**Neden:** {row['Risk Nedenleri']}")
                    else:
                        st.success("Riskli mağaza yok! 🎉")
                
                with tabs[3]:
                    st.subheader("🚬 Sigara Açığı Olan Mağazalar")
                    sigara_df = region_df[region_df['Sigara'] > 0].sort_values('Sigara', ascending=False)
                    if len(sigara_df) > 0:
                        st.error(f"⚠️ {len(sigara_df)} mağazada sigara açığı var!")
                        for _, row in sigara_df.iterrows():
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**: {row['Sigara']:.0f} ürün sigara açığı")
                    else:
                        st.success("Sigara açığı olan mağaza yok! 🎉")
                
                with tabs[4]:
                    st.subheader("📊 Tüm Detaylar")
                    st.dataframe(region_df, use_container_width=True, hide_index=True)
                
                with tabs[5]:
                    st.subheader("📥 SM Raporu İndir")
                    
                    excel_data = create_region_excel_report(region_df, df, kasa_kodlari, params)
                    
                    st.download_button(
                        label=f"📥 {display_sm} Özet Raporu (Excel)",
                        data=excel_data,
                        file_name=f"SM_OZET_{display_sm}_{params.get('donem', '')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

elif uploaded_file is not None:
    try:
        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names
        
        best_sheet = None
        max_cols = 0
        
        for sheet in sheet_names:
            temp_df = pd.read_excel(uploaded_file, sheet_name=sheet, nrows=5)
            if len(temp_df.columns) > max_cols:
                max_cols = len(temp_df.columns)
                best_sheet = sheet
        
        df_raw = pd.read_excel(uploaded_file, sheet_name=best_sheet)
        st.success(f"✅ {len(df_raw)} satır, {len(df_raw.columns)} sütun ({best_sheet})")
        
        # ===== ARKA PLANDA SUPABASE'E KAYIT =====
        with st.spinner("Veritabanına kaydediliyor..."):
            try:
                inserted, skipped, result_info = save_to_supabase(df_raw)
                if inserted > 0:
                    st.info(f"💾 {inserted:,} kayıt eklendi | ⏭️ {skipped} envanter zaten mevcut")
                elif skipped > 0:
                    st.info(f"⏭️ Tüm envanterler zaten mevcut ({skipped} envanter)")
            except Exception as e:
                # Supabase hatası analizi engellemesin
                st.warning(f"⚠️ Veritabanı kaydı atlandı: {str(e)[:50]}")
        
        df = analyze_inventory(df_raw)
        
        # Mağaza bilgisi
        if 'Mağaza Kodu' in df.columns:
            magazalar = df['Mağaza Kodu'].dropna().unique().tolist()
            # Mağaza kod-isim eşleştirmesi
            magaza_isimleri = {}
            for mag in magazalar:
                isim = df[df['Mağaza Kodu'] == mag]['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df.columns else ''
                magaza_isimleri[mag] = f"{mag} - {isim}" if isim else str(mag)
        else:
            magazalar = ['MAGAZA']
            df['Mağaza Kodu'] = 'MAGAZA'
            magaza_isimleri = {'MAGAZA': 'MAGAZA'}
        
        params = {
            'donem': str(df['Envanter Dönemi'].iloc[0]) if 'Envanter Dönemi' in df.columns else '',
            'tarih': str(df['Envanter Tarihi'].iloc[0])[:10] if 'Envanter Tarihi' in df.columns else '',
        }
        
        # Kasa aktivitesi kodlarını yükle
        kasa_kodlari = load_kasa_activity_codes()
        
        # ========== BÖLGE ÖZETİ MODU ==========
        if analysis_mode == "🌍 Bölge Özeti":
            st.subheader(f"🌍 Bölge Özeti - {len(magazalar)} Mağaza")
            
            with st.spinner("Tüm mağazalar analiz ediliyor..."):
                region_df = analyze_region(df, kasa_kodlari)
            
            if len(region_df) == 0:
                st.warning("Analiz edilecek mağaza bulunamadı!")
            else:
                # Bölge toplamları
                toplam_satis = region_df['Satış'].sum()
                toplam_fark = region_df['Fark'].sum()  # Fark + Kısmi
                toplam_fire = region_df['Fire'].sum()
                toplam_acik = region_df['Toplam Açık'].sum()  # Fark + Fire
                toplam_gun = region_df['Gün'].sum()
                
                # Oranlar
                fark_oran = abs(toplam_fark) / toplam_satis * 100 if toplam_satis > 0 else 0
                fire_oran = abs(toplam_fire) / toplam_satis * 100 if toplam_satis > 0 else 0
                toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
                gunluk_fark = toplam_fark / toplam_gun if toplam_gun > 0 else 0
                gunluk_fire = toplam_fire / toplam_gun if toplam_gun > 0 else 0
                
                # Risk dağılımı
                kritik_sayisi = len(region_df[region_df['Risk'].str.contains('KRİTİK')])
                riskli_sayisi = len(region_df[region_df['Risk'].str.contains('RİSKLİ')])
                dikkat_sayisi = len(region_df[region_df['Risk'].str.contains('DİKKAT')])
                temiz_sayisi = len(region_df[region_df['Risk'].str.contains('TEMİZ')])
                
                # Üst metrikler
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("💰 Toplam Satış", f"{toplam_satis/1_000_000:.1f}M TL")
                with col2:
                    st.metric("📉 Fark", f"{toplam_fark:,.0f} TL", f"%{fark_oran:.2f} | Günlük: {gunluk_fark:,.0f}₺")
                with col3:
                    st.metric("🔥 Fire", f"{toplam_fire:,.0f} TL", f"%{fire_oran:.2f} | Günlük: {gunluk_fire:,.0f}₺")
                with col4:
                    st.metric("📊 Toplam", f"{toplam_acik:,.0f} TL", f"%{toplam_oran:.2f}")
                
                # Risk dağılımı
                st.markdown("### 📊 Risk Dağılımı")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    if kritik_sayisi > 0:
                        st.markdown(f'<div class="risk-kritik">🔴 KRİTİK: {kritik_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🔴 KRİTİK", kritik_sayisi)
                with col2:
                    if riskli_sayisi > 0:
                        st.markdown(f'<div class="risk-riskli">🟠 RİSKLİ: {riskli_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🟠 RİSKLİ", riskli_sayisi)
                with col3:
                    if dikkat_sayisi > 0:
                        st.markdown(f'<div class="risk-dikkat">🟡 DİKKAT: {dikkat_sayisi}</div>', unsafe_allow_html=True)
                    else:
                        st.metric("🟡 DİKKAT", dikkat_sayisi)
                with col4:
                    st.markdown(f'<div class="risk-temiz">🟢 TEMİZ: {temiz_sayisi}</div>', unsafe_allow_html=True)
                
                # Sekmeler
                tabs = st.tabs(["📋 Sıralama", "🔴 Kritik", "🟠 Riskli", "🚬 Sigara", "📊 Detay", "📥 İndir"])
                
                with tabs[0]:
                    st.subheader("📋 Mağaza Sıralaması (Risk Puanına Göre)")
                    
                    # Başlık satırı
                    cols = st.columns([0.4, 0.8, 1.3, 1.2, 0.9, 0.7, 0.9, 0.7, 0.6, 0.6, 0.4, 0.5, 0.8])
                    cols[0].markdown("**📥**")
                    cols[1].markdown("**Kod**")
                    cols[2].markdown("**Mağaza Adı**")
                    cols[3].markdown("**BS**")
                    cols[4].markdown("**Fark**")
                    cols[5].markdown("**Günlük**")
                    cols[6].markdown("**Fire**")
                    cols[7].markdown("**Günlük**")
                    cols[8].markdown("**Kayıp%**")
                    cols[9].markdown("**Fire%**")
                    cols[10].markdown("**🚬**")
                    cols[11].markdown("**Risk**")
                    cols[12].markdown("**Seviye**")
                    
                    st.markdown("---")
                    
                    # Veri satırları
                    for idx, (_, row) in enumerate(region_df.iterrows()):
                        cols = st.columns([0.4, 0.8, 1.3, 1.2, 0.9, 0.7, 0.9, 0.7, 0.6, 0.6, 0.4, 0.5, 0.8])
                        
                        # Mağaza verisini al ve tam rapor oluştur
                        mag_kod = row['Mağaza Kodu']
                        df_mag = df[df['Mağaza Kodu'] == mag_kod].copy()
                        mag_adi = row['Mağaza Adı']
                        
                        # Analizleri yap
                        int_df = detect_internal_theft(df_mag)
                        chr_df = detect_chronic_products(df_mag)
                        chr_fire_df = detect_chronic_fire(df_mag)
                        cig_df = detect_cigarette_shortage(df_mag)
                        ext_df = detect_external_theft(df_mag)
                        fam_df = find_product_families(df_mag)
                        fire_df = detect_fire_manipulation(df_mag)
                        kasa_df, kasa_sum = check_kasa_activity_products(df_mag, kasa_kodlari)
                        
                        int_codes = set(int_df['Malzeme Kodu'].astype(str).tolist()) if len(int_df) > 0 else set()
                        chr_codes = set(chr_df['Malzeme Kodu'].astype(str).tolist()) if len(chr_df) > 0 else set()
                        
                        t20_df = create_top_20_risky(df_mag, int_codes, chr_codes, set())
                        exec_c, grp_s = generate_executive_summary(df_mag, kasa_df, kasa_sum)
                        
                        # Tam rapor oluştur
                        report_data = create_excel_report(
                            df_mag, int_df, chr_df, chr_fire_df, cig_df,
                            ext_df, fam_df, fire_df, kasa_df, t20_df,
                            exec_c, grp_s, mag_kod, mag_adi, params
                        )
                        
                        mag_adi_clean = mag_adi.replace(' ', '_').replace('/', '_')[:30] if mag_adi else ''
                        
                        with cols[0]:
                            st.download_button("📥", data=report_data, 
                                file_name=f"{mag_kod}_{mag_adi_clean}_Risk_Raporu.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_{idx}")
                        cols[1].write(f"{row['Mağaza Kodu']}")
                        cols[2].write(f"{row['Mağaza Adı'][:18] if row['Mağaza Adı'] else '-'}")
                        cols[3].write(f"{row['BS'][:12] if row['BS'] else '-'}")
                        cols[4].write(f"{row['Fark']:,.0f}")
                        cols[5].write(f"{row['Günlük Fark']:,.0f}")
                        cols[6].write(f"{row['Fire']:,.0f}")
                        cols[7].write(f"{row['Günlük Fire']:,.0f}")
                        cols[8].write(f"%{row['Toplam %']:.1f}")
                        cols[9].write(f"%{row['Fire %']:.1f}")
                        cols[10].write(f"{row['Sigara']}" if row['Sigara'] > 0 else "-")
                        cols[11].write(f"{row['Risk Puan']:.0f}")
                        cols[12].write(row['Risk'])
                
                with tabs[1]:
                    st.subheader("🔴 Kritik Mağazalar")
                    kritik_df = region_df[region_df['Risk'].str.contains('KRİTİK')]
                    if len(kritik_df) > 0:
                        for _, row in kritik_df.iterrows():
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**\n\n"
                                    f"Kayıp: %{row['Toplam %']:.1f} | Fark: {row['Fark']:,.0f} TL\n\n"
                                    f"**Neden:** {row['Risk Nedenleri']}")
                    else:
                        st.success("Kritik mağaza yok! 🎉")
                
                with tabs[2]:
                    st.subheader("🟠 Riskli Mağazalar")
                    riskli_df = region_df[region_df['Risk'].str.contains('RİSKLİ')]
                    if len(riskli_df) > 0:
                        for _, row in riskli_df.iterrows():
                            st.warning(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**\n\n"
                                      f"Kayıp: %{row['Toplam %']:.1f} | Fark: {row['Fark']:,.0f} TL\n\n"
                                      f"**Neden:** {row['Risk Nedenleri']}")
                    else:
                        st.success("Riskli mağaza yok! 🎉")
                
                with tabs[3]:
                    st.subheader("🚬 Sigara Açığı Olan Mağazalar")
                    sigara_df = region_df[region_df['Sigara'] > 0].sort_values('Sigara', ascending=False)
                    if len(sigara_df) > 0:
                        st.error(f"⚠️ {len(sigara_df)} mağazada sigara açığı var!")
                        for _, row in sigara_df.iterrows():
                            st.error(f"**{row['Mağaza Kodu']} - {row['Mağaza Adı']}**: {row['Sigara']} ürün sigara açığı")
                    else:
                        st.success("Sigara açığı olan mağaza yok! 🎉")
                
                with tabs[4]:
                    st.subheader("📊 Tüm Detaylar")
                    st.dataframe(region_df, use_container_width=True, hide_index=True)
                
                with tabs[5]:
                    st.subheader("📥 Bölge Raporu İndir")
                    
                    excel_data = create_region_excel_report(region_df, df, kasa_kodlari, params)
                    
                    st.download_button(
                        label="📥 Bölge Özet Raporu (Excel)",
                        data=excel_data,
                        file_name=f"BOLGE_OZET_{params.get('donem', '')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        # ========== TEK MAĞAZA MODU ==========
        else:
            # Mağaza seçimi
            if len(magazalar) > 1:
                # Kod + isim listesi oluştur
                magaza_options = [magaza_isimleri[m] for m in magazalar]
                selected_option = st.selectbox("🏪 Mağaza Seçin", magaza_options)
                # Seçilen option'dan kodu çıkar
                selected_str = selected_option.split(" - ")[0]
                # Orijinal tipte bul
                selected = None
                for m in magazalar:
                    if str(m) == selected_str:
                        selected = m
                        break
                if selected is None:
                    selected = magazalar[0]
                df_display = df[df['Mağaza Kodu'] == selected].copy()
                magaza_adi = df_display['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df_display.columns and len(df_display) > 0 else ''
            else:
                selected = magazalar[0]
                df_display = df.copy()
                magaza_adi = df['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df.columns and len(df) > 0 else ''
        
            # Kasa aktivitesi kodlarını yükle
            kasa_kodlari = load_kasa_activity_codes()
        
            # Analizler
            internal_df = detect_internal_theft(df_display)
            chronic_df = detect_chronic_products(df_display)
            chronic_fire_df = detect_chronic_fire(df_display)
            cigarette_df = detect_cigarette_shortage(df_display)
            external_df = detect_external_theft(df_display)
            family_df = find_product_families(df_display)
            fire_manip_df = detect_fire_manipulation(df_display)
            kasa_activity_df, kasa_summary = check_kasa_activity_products(df_display, kasa_kodlari)
            exec_comments, group_stats = generate_executive_summary(df_display, kasa_activity_df, kasa_summary)
        
            internal_codes = set(internal_df['Malzeme Kodu'].astype(str).tolist()) if len(internal_df) > 0 else set()
            chronic_codes = set(chronic_df['Malzeme Kodu'].astype(str).tolist()) if len(chronic_df) > 0 else set()
        
            # Aile dengelenmişlerini bul
            family_balanced_codes = set()
            if len(family_df) > 0:
                balanced_families = family_df[family_df['Sonuç'].str.contains('KARIŞIKLIK', na=False)]
                # Bu ailelerdeki ürünleri bul
        
            top20_df = create_top_20_risky(df_display, internal_codes, chronic_codes, family_balanced_codes)
        
            risk_seviyesi, risk_class = calculate_store_risk(df_display, internal_df, chronic_df, cigarette_df)
        
            st.markdown("---")
        
            # Metrikler hesapla
            toplam_satis = df_display['Satış Tutarı'].sum()
            fark_tutari = df_display['Fark Tutarı'].fillna(0).sum()
            kismi_tutari = df_display['Kısmi Envanter Tutarı'].fillna(0).sum()
            fire_tutari = df_display['Fire Tutarı'].fillna(0).sum()
            
            fark = fark_tutari + kismi_tutari
            toplam_acik = fark + fire_tutari
            
            fark_oran = abs(fark) / toplam_satis * 100 if toplam_satis > 0 else 0
            fire_oran = abs(fire_tutari) / toplam_satis * 100 if toplam_satis > 0 else 0
            toplam_oran = abs(toplam_acik) / toplam_satis * 100 if toplam_satis > 0 else 0
            
            # Gün hesabı
            gun_sayisi = 1
            try:
                if 'Envanter Tarihi' in df_display.columns and 'Envanter Başlangıç Tarihi' in df_display.columns:
                    env_tarihi = pd.to_datetime(df_display['Envanter Tarihi'].iloc[0])
                    env_baslangic = pd.to_datetime(df_display['Envanter Başlangıç Tarihi'].iloc[0])
                    gun_sayisi = (env_tarihi - env_baslangic).days
                    if gun_sayisi <= 0:
                        gun_sayisi = 1
            except:
                gun_sayisi = 1
            
            gunluk_fark = fark / gun_sayisi
            gunluk_fire = fire_tutari / gun_sayisi
        
            # Metrikler - Üst
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.markdown(f'<div class="{risk_class}"><b>RİSK</b><br/><h2>{risk_seviyesi}</h2></div>', unsafe_allow_html=True)
            with col2:
                st.metric("💰 Satış", f"{toplam_satis:,.0f} TL")
            with col3:
                st.metric("📉 Fark", f"{fark:,.0f} TL", f"%{fark_oran:.2f} | Günlük: {gunluk_fark:,.0f}₺")
            with col4:
                st.metric("🔥 Fire", f"{fire_tutari:,.0f} TL", f"%{fire_oran:.2f} | Günlük: {gunluk_fire:,.0f}₺")
            with col5:
                st.metric("📊 Toplam", f"{toplam_acik:,.0f} TL", f"%{toplam_oran:.2f}")
        
            # Metrikler - Alt
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("🔒 İç Hırs.", f"{len(internal_df)}")
            with col2:
                st.metric("🔄 Kr.Açık", f"{len(chronic_df)}")
            with col3:
                st.metric("🔥 Kr.Fire", f"{len(chronic_fire_df)}")
            with col4:
                # Sigara açığı - toplam bazlı
                sigara_acik = 0
                if len(cigarette_df) > 0 and 'Ürün Toplam' in cigarette_df.columns:
                    son_satir = cigarette_df.iloc[-1]
                    if son_satir['Malzeme Kodu'] == '*** TOPLAM ***':
                        sigara_acik = abs(son_satir['Ürün Toplam'])
                
                if sigara_acik > 0:
                    st.metric("🚬 SİGARA", f"{sigara_acik:.0f}", delta="RİSK!", delta_color="inverse")
                else:
                    st.metric("🚬 Sigara", "0")
            with col5:
                if kasa_summary['toplam_adet'] > 0:
                    st.metric("💰 10 TL", f"+{kasa_summary['toplam_adet']:.0f} / {kasa_summary['toplam_tutar']:,.0f}₺", delta="FAZLA!", delta_color="inverse")
                elif kasa_summary['toplam_adet'] < 0:
                    st.metric("💰 10 TL", f"{kasa_summary['toplam_adet']:.0f} / {kasa_summary['toplam_tutar']:,.0f}₺", delta="AÇIK", delta_color="normal")
                else:
                    st.metric("💰 10 TL", "0")
        
            # Yönetici Özeti
            if exec_comments:
                with st.expander("📋 Yönetici Özeti", expanded=True):
                    for comment in exec_comments[:5]:
                        st.markdown(comment)
        
            st.markdown("---")
        
            # Sekmeler
            tabs = st.tabs(["🚨 Riskli 20", "🔒 İç Hırs.", "🔄 Kr.Açık", "🔥 Kr.Fire", "🔥 Fire Man.", "🚬 Sigara", "💰 10 TL Akt.", "📥 İndir"])
        
            with tabs[0]:
                st.subheader("🚨 En Riskli 20 Ürün")
                if len(top20_df) > 0:
                    st.dataframe(top20_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Riskli ürün yok!")
        
            with tabs[1]:
                st.subheader("🔒 İç Hırsızlık (≥100TL)")
                st.caption("Fark büyüdükçe risk AZALIR, eşitse EN YÜKSEK")
                if len(internal_df) > 0:
                    st.dataframe(internal_df, use_container_width=True, hide_index=True)
                else:
                    st.success("İç hırsızlık riski yok!")
        
            with tabs[2]:
                st.subheader("🔄 Kronik Açık")
                st.caption("Her iki dönemde de Fark < 0")
                if len(chronic_df) > 0:
                    st.dataframe(chronic_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Kronik açık yok!")
        
            with tabs[3]:
                st.subheader("🔥 Kronik Fire")
                st.caption("Her iki dönemde de fire kaydı var")
                if len(chronic_fire_df) > 0:
                    st.dataframe(chronic_fire_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Kronik fire yok!")
        
            with tabs[4]:
                st.subheader("🔥 Fire Manipülasyonu")
                st.caption("Fire var ama Fark+Kısmi > 0")
                if len(fire_manip_df) > 0:
                    st.dataframe(fire_manip_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Fire manipülasyonu yok!")
        
            with tabs[5]:
                st.subheader("🚬 Sigara Açığı")
                if len(cigarette_df) > 0:
                    st.error("⚠️ Sigarada açık = HIRSIZLIK BELİRTİSİ")
                    st.dataframe(cigarette_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Sigara açığı yok!")
        
            with tabs[6]:
                st.subheader("💰 10 TL Aktivitesi Ürünleri")
            
                if kasa_summary['toplam_adet'] != 0:
                    if kasa_summary['toplam_adet'] > 0:
                        st.error(f"⚠️ NET +{kasa_summary['toplam_adet']:.0f} adet / {kasa_summary['toplam_tutar']:,.0f} TL FAZLA - Gerçek açığı gizliyor olabilir!")
                    else:
                        st.warning(f"📉 NET {kasa_summary['toplam_adet']:.0f} adet / {kasa_summary['toplam_tutar']:,.0f} TL AÇIK")
            
                if len(kasa_activity_df) > 0:
                    st.dataframe(kasa_activity_df, use_container_width=True, hide_index=True)
                else:
                    st.success("Kasa aktivitesi ürünlerinde sorun yok!")
        
            with tabs[7]:
                st.subheader("📥 Rapor İndir")
            
                excel_output = create_excel_report(
                    df_display, internal_df, chronic_df, chronic_fire_df, cigarette_df,
                    external_df, family_df, fire_manip_df, kasa_activity_df, top20_df,
                    exec_comments, group_stats, selected, magaza_adi, params
                )
                
                mag_adi_clean = magaza_adi.replace(' ', '_').replace('/', '_')[:30] if magaza_adi else ''
            
                st.download_button(
                    label=f"📥 {selected} Raporu İndir",
                    data=excel_output,
                    file_name=f"{selected}_{mag_adi_clean}_Risk_Raporu.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
                if len(magazalar) > 1:
                    st.markdown("---")
                    if st.button("🗜️ Tüm Mağazaları Hazırla (ZIP)"):
                        with st.spinner("Raporlar hazırlanıyor..."):
                            zip_buffer = BytesIO()
                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                                for mag in magazalar:
                                    df_mag = df[df['Mağaza Kodu'] == mag].copy()
                                    mag_adi = df_mag['Mağaza Adı'].iloc[0] if 'Mağaza Adı' in df_mag.columns and len(df_mag) > 0 else ''
                                
                                    int_df = detect_internal_theft(df_mag)
                                    chr_df = detect_chronic_products(df_mag)
                                    chr_fire_df = detect_chronic_fire(df_mag)
                                    cig_df = detect_cigarette_shortage(df_mag)
                                    ext_df = detect_external_theft(df_mag)
                                    fam_df = find_product_families(df_mag)
                                    fire_df = detect_fire_manipulation(df_mag)
                                    kasa_df, kasa_sum = check_kasa_activity_products(df_mag, kasa_kodlari)
                                
                                    int_codes = set(int_df['Malzeme Kodu'].astype(str).tolist()) if len(int_df) > 0 else set()
                                    chr_codes = set(chr_df['Malzeme Kodu'].astype(str).tolist()) if len(chr_df) > 0 else set()
                                
                                    t20_df = create_top_20_risky(df_mag, int_codes, chr_codes, set())
                                    exec_c, grp_s = generate_executive_summary(df_mag, kasa_df, kasa_sum)
                                
                                    excel_data = create_excel_report(
                                        df_mag, int_df, chr_df, chr_fire_df, cig_df,
                                        ext_df, fam_df, fire_df, kasa_df, t20_df,
                                        exec_c, grp_s, mag, mag_adi, params
                                    )
                                
                                    zf.writestr(f"{mag}_Risk_Raporu.xlsx", excel_data.getvalue())
                        
                            zip_buffer.seek(0)
                            st.download_button(
                                label=f"📥 {len(magazalar)} Mağaza ZIP İndir",
                                data=zip_buffer,
                                file_name="Tum_Magazalar_Rapor.zip",
                                mime="application/zip"
                            )
    
    except Exception as e:
        st.error(f"Hata: {str(e)}")
        st.exception(e)

else:
    if analysis_mode != "👔 SM Özet":
        st.info("👆 Excel dosyası yükleyin")
